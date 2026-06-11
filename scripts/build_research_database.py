from __future__ import annotations

import hashlib
import re
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "数据库"
INPUT_XLSX = DATA_DIR / "5.20完善数据集.xlsx"
INPUT_BIB = DATA_DIR / "Assembly Method.bib"
OUTPUT_XLSX = DATA_DIR / "5.20完善数据集_research_db_v0.1.xlsx"


def parse_bibtex(path: Path) -> pd.DataFrame:
    text = path.read_text(encoding="utf-8", errors="replace")
    entries = []
    pattern = re.compile(r"@(?P<type>\w+)\s*\{\s*(?P<key>[^,]+),(?P<body>.*?)(?=\n@\w+\s*\{|\Z)", re.S)

    for match in pattern.finditer(text):
        body = match.group("body")
        fields = parse_bib_fields(body)

        key = match.group("key").strip()
        entries.append(
            {
                "Source_ID": f"SRC-{len(entries) + 1:04d}",
                "BibTeX_Key": key,
                "Source_Type": match.group("type").lower(),
                "Title": fields.get("title", ""),
                "Authors": fields.get("author", ""),
                "Year": fields.get("date", fields.get("year", ""))[:4],
                "Journal": fields.get("journaltitle", fields.get("journal", "")),
                "DOI_or_URL": fields.get("doi", fields.get("url", "")),
                "PMID": fields.get("eprint", "") if fields.get("eprinttype", "").lower() == "pubmed" else "",
                "Source_Domain": infer_source_domain(key, fields),
                "Search_Query": "",
                "Why_Relevant": infer_relevance(fields),
                "Access_Date": fields.get("urldate", ""),
                "Screening_Status": "candidate",
                "Exclusion_Reason": "",
                "Notes": "",
            }
        )
    return pd.DataFrame(entries)


def parse_bib_fields(body: str) -> dict[str, str]:
    fields: dict[str, str] = {}
    i = 0
    n = len(body)
    while i < n:
        while i < n and (body[i].isspace() or body[i] == ","):
            i += 1
        name_start = i
        while i < n and (body[i].isalnum() or body[i] in "_-"):
            i += 1
        if i == name_start:
            break
        name = body[name_start:i].lower()
        while i < n and body[i].isspace():
            i += 1
        if i >= n or body[i] != "=":
            continue
        i += 1
        while i < n and body[i].isspace():
            i += 1
        if i >= n:
            break

        if body[i] == "{":
            depth = 0
            value_start = i + 1
            while i < n:
                if body[i] == "{":
                    depth += 1
                elif body[i] == "}":
                    depth -= 1
                    if depth == 0:
                        value = body[value_start:i]
                        i += 1
                        break
                i += 1
            else:
                value = body[value_start:]
        elif body[i] == '"':
            i += 1
            value_start = i
            while i < n and body[i] != '"':
                i += 1
            value = body[value_start:i]
            i += 1
        else:
            value_start = i
            while i < n and body[i] not in ",\n":
                i += 1
            value = body[value_start:i]

        value = value.replace("{{", "").replace("}}", "")
        value = value.replace("{", "").replace("}", "")
        fields[name] = re.sub(r"\s+", " ", value).strip()
    return fields


def infer_source_domain(key: str, fields: dict[str, str]) -> str:
    hay = " ".join([key, fields.get("title", ""), fields.get("abstract", ""), fields.get("journaltitle", "")]).lower()
    rules = [
        ("Bennu", ["bennu", "osiris"]),
        ("Ryugu", ["ryugu", "hayabusa"]),
        ("Combustion", ["combustion", "flame", "pyrolysis", "pah", "polycyclic aromatic"]),
        ("Meteorite_IOM", ["meteorite", "murchison", "insoluble organic matter", " iom "]),
        ("Assembly_Theory", ["assembly theory", "assembly index", "molecular assembly"]),
        ("Machine_Learning", ["machine learning", "neural", "transformer"]),
        ("Biosignature_Framework", ["biosignature", "life detection", "darwinism"]),
        ("Molecular_Complexity", ["molecular complexity", "algorithmic complexity", "shannon", "entropy"]),
        ("Natural_Product", ["natural product", "biomarker", "metabolite"]),
        ("Prebiotic", ["prebiotic", "abiotic", "origin of life", "interstellar", "ice"]),
    ]
    for domain, needles in rules:
        if any(n in hay for n in needles):
            return domain
    return "Unclassified"


def infer_relevance(fields: dict[str, str]) -> str:
    hay = " ".join([fields.get("title", ""), fields.get("abstract", "")]).lower()
    if "assembly" in hay:
        return "Assembly theory / MA method or critique"
    if any(x in hay for x in ["bennu", "ryugu", "meteorite", "murchison"]):
        return "Extraterrestrial organic matter source evidence"
    if any(x in hay for x in ["combustion", "pyrolysis", "polycyclic aromatic", "flame"]):
        return "Abiotic high-complexity molecule source evidence"
    if any(x in hay for x in ["biosignature", "life detection", "darwinism"]):
        return "Biosignature interpretation framework"
    if "machine learning" in hay:
        return "Future ML feature/model reference"
    return ""


def normalize_origin(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def origin_class(origin: str) -> str:
    low = origin.lower()
    if low in {"combustion", "pahmethane pyrolysis"}:
        return "Abiotic_Strong"
    if low in {"bennu", "ryugu", "prebio"}:
        return "Abiotic_Strong"
    if low in {"biomarker", "siderophore"}:
        return "Bio_Strong_Candidate"
    if low == "peptide":
        return "Bio_Ambiguous"
    return "Unclassified"


def origin_label(origin: str) -> str:
    cls = origin_class(origin)
    if cls.startswith("Abiotic"):
        return "abiotic"
    if cls.startswith("Bio"):
        return "biological"
    return "ambiguous"


def label_strength(origin: str) -> str:
    cls = origin_class(origin)
    if cls == "Abiotic_Strong":
        return "strong"
    if cls == "Bio_Strong_Candidate":
        return "candidate_strong"
    if cls == "Bio_Ambiguous":
        return "weak"
    return "unchecked"


def label_basis(origin: str) -> str:
    low = origin.lower()
    if low in {"bennu", "ryugu"}:
        return "extraterrestrial_sample_context"
    if low == "prebio":
        return "abiotic_experiment_or_prebiotic_context"
    if low in {"combustion", "pahmethane pyrolysis"}:
        return "abiotic_experimental_or_process_context"
    if low in {"biomarker", "siderophore"}:
        return "biosynthetic_or_biomarker_literature_claim"
    if low == "peptide":
        return "biological_class_annotation"
    return "not_checked"


def graph_level(row: pd.Series) -> str:
    if pd.notna(row.get("Canonical_SMILES")) or pd.notna(row.get("Isomeric_SMILES")):
        return "G3_2D_graph_available"
    if pd.notna(row.get("Formula")):
        return "G1_formula_only_or_unresolved"
    return "G0_no_single_molecular_graph"


def graph_ambiguity(row: pd.Series) -> str:
    if graph_level(row).startswith("G3"):
        if pd.isna(row.get("Isomeric_SMILES")):
            return "stereochemistry_missing_or_not_applicable"
        return "none_recorded"
    if pd.notna(row.get("Formula")):
        return "isomer_uncertain_formula_only"
    return "structure_missing"


def ma_use(row: pd.Series) -> str:
    if pd.notna(row.get("MA")) and graph_level(row).startswith("G3"):
        return "MA_primary"
    if pd.notna(row.get("MA")):
        return "MA_sensitivity_or_requires_mol_audit"
    return "exclude_until_MA_computed"


def confidence(row: pd.Series) -> str:
    origin = normalize_origin(row.get("Origin"))
    if graph_level(row).startswith("G3") and label_strength(origin) in {"strong", "candidate_strong"}:
        return "B_requires_source_link"
    if pd.notna(row.get("MA")):
        return "C_requires_structure_or_source_audit"
    return "D_unusable_for_analysis_now"


def source_domain(origin: str) -> str:
    mapping = {
        "Bennu": "Bennu",
        "Ryugu": "Ryugu",
        "Combustion": "Combustion",
        "PAHmethane pyrolysis": "Combustion",
        "Prebio": "Prebiotic",
        "Biomarker": "Natural_Product",
        "siderophore": "Natural_Product",
        "peptide": "Peptide",
    }
    return mapping.get(origin, "Unclassified")


def stable_id(prefix: str, idx: int) -> str:
    return f"{prefix}-{idx:05d}"


def sha256_text(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8", errors="replace")).hexdigest()


def build_tables() -> dict[str, pd.DataFrame]:
    xl = pd.ExcelFile(INPUT_XLSX)
    original = {name: pd.read_excel(INPUT_XLSX, sheet_name=name) for name in xl.sheet_names}
    all_df = original["All"].copy()
    bib_df = parse_bibtex(INPUT_BIB)

    all_df.insert(0, "Original_Row", range(2, len(all_df) + 2))
    all_df.insert(0, "Record_ID", [stable_id("REC", i + 1) for i in range(len(all_df))])

    evidence = pd.DataFrame(
        {
            "Record_ID": all_df["Record_ID"],
            "Source_ID": "",
            "Source_Domain": all_df["Origin"].map(lambda x: source_domain(normalize_origin(x))),
            "Original_Name_in_Source": all_df["Original_Name"],
            "Reported_Formula": all_df.get("Formula", ""),
            "Reported_SMILES_or_Structure": all_df.get("Canonical_SMILES", ""),
            "Reported_Context": all_df["Origin"].map(normalize_origin),
            "Reported_Origin_Label": all_df["Origin"].map(lambda x: origin_label(normalize_origin(x))),
            "Evidence_Location": "",
            "Extraction_Method": "legacy_excel_import",
            "Extractor": "LiYangu/Codex legacy migration",
            "Extraction_Date": str(date.today()),
            "Evidence_Confidence": all_df.apply(confidence, axis=1),
            "Notes": "Migrated from 5.20完善数据集.xlsx; source paper/table still needs manual linking.",
        }
    )

    curated = pd.DataFrame()
    curated["Curated_ID"] = [stable_id("MOL", i + 1) for i in range(len(all_df))]
    curated["Record_IDs"] = all_df["Record_ID"]
    curated["Preferred_Name"] = all_df["Cleaned_Query"].where(all_df["Cleaned_Query"].notna(), all_df["Original_Name"])
    curated["Original_Name"] = all_df["Original_Name"]
    curated["IUPAC_Name"] = all_df.get("IUPAC_Name", "")
    curated["Formula"] = all_df.get("Formula", "")
    curated["Molecular_Weight"] = all_df.get("Molecular_Weight", "")
    curated["Exact_Mass"] = all_df.get("Exact_Mass", "")
    curated["Canonical_SMILES"] = all_df.get("Canonical_SMILES", "")
    curated["Isomeric_SMILES"] = all_df.get("Isomeric_SMILES", "")
    curated["InChIKey"] = all_df.get("InChIKey", "")
    curated["CID"] = all_df.get("CID", "")
    curated["Origin"] = all_df["Origin"]
    curated["Type"] = all_df.get("Type", "")
    curated["MA"] = all_df.get("MA", "")
    curated["Source_Domain"] = all_df["Origin"].map(lambda x: source_domain(normalize_origin(x)))
    curated["Origin_Class"] = all_df["Origin"].map(lambda x: origin_class(normalize_origin(x)))
    curated["Origin_Label"] = all_df["Origin"].map(lambda x: origin_label(normalize_origin(x)))
    curated["Label_Strength"] = all_df["Origin"].map(lambda x: label_strength(normalize_origin(x)))
    curated["Label_Basis"] = all_df["Origin"].map(lambda x: label_basis(normalize_origin(x)))
    curated["Nonbiotic_Report_Status"] = all_df["Origin"].map(
        lambda x: "not_applicable_abiotic" if origin_label(normalize_origin(x)) == "abiotic" else "not_checked"
    )
    curated["Mol_Available"] = all_df.apply(lambda r: "yes" if graph_level(r).startswith("G3") else "unknown_or_no", axis=1)
    curated["Mol_Source"] = all_df.apply(lambda r: "legacy_smiles_or_mol" if graph_level(r).startswith("G3") else "needs_mol_file_audit", axis=1)
    curated["Mol_File"] = ""
    curated["Mol_File_SHA256"] = ""
    curated["Graph_Level"] = all_df.apply(graph_level, axis=1)
    curated["Graph_Ambiguity"] = all_df.apply(graph_ambiguity, axis=1)
    curated["MA_Use"] = all_df.apply(ma_use, axis=1)
    curated["Structure_Validation_Status"] = all_df.apply(
        lambda r: "structure_fields_present_needs_audit" if graph_level(r).startswith("G3") else "needs_mol_or_structure_resolution",
        axis=1,
    )
    curated["Source_Validation_Status"] = "needs_source_registry_link"
    curated["Confidence"] = all_df.apply(confidence, axis=1)
    duplicate_inchikey = all_df["InChIKey"].notna() & all_df["InChIKey"].duplicated(keep=False)
    curated["Conflict_Flag"] = duplicate_inchikey.map(lambda x: "possible_duplicate_or_multi_source" if x else "no")
    curated["Conflict_Description"] = ""
    curated["Curation_Note"] = ""

    ml = curated.copy()
    ml["Train_Eligible"] = ml.apply(
        lambda r: "provisional_after_source_verification"
        if r["Graph_Level"].startswith("G3")
        and r["Label_Strength"] in {"strong", "candidate_strong"}
        and pd.notna(r["MA"])
        else "no",
        axis=1,
    )
    ml["Reason_Excluded"] = ml.apply(reason_excluded, axis=1)
    ml["Dataset_Split"] = ""
    ml["Scaffold_ID"] = ""
    ml["Leakage_Group"] = ml["InChIKey"].fillna("").astype(str).str[:14]
    ml = ml[
        [
            "Curated_ID",
            "Canonical_SMILES",
            "Isomeric_SMILES",
            "InChIKey",
            "Mol_File",
            "Origin_Label",
            "Label_Strength",
            "Label_Basis",
            "Source_Domain",
            "Origin_Class",
            "MA",
            "Molecular_Weight",
            "Exact_Mass",
            "Graph_Level",
            "Graph_Ambiguity",
            "Train_Eligible",
            "Reason_Excluded",
            "Dataset_Split",
            "Scaffold_ID",
            "Leakage_Group",
            "Confidence",
        ]
    ]

    high_abiotic = curated[
        (curated["Origin_Label"] == "abiotic") & (pd.to_numeric(curated["MA"], errors="coerce") >= 15)
    ].copy()
    high_abiotic["Why_it_challenges_MA_threshold"] = "Abiotic-labeled molecule with MA >= 15; requires source verification before publication."

    low_bio = curated[
        (curated["Origin_Label"] == "biological") & (pd.to_numeric(curated["MA"], errors="coerce") < 15)
    ].copy()
    low_bio["Why_it_challenges_low_MA_exclusion"] = "Biological-labeled candidate with MA < 15; requires nonbiotic-report search before strong claim."

    checks = make_checks(all_df, curated, bib_df, high_abiotic, low_bio)
    dictionary = make_dictionary()
    readme = make_readme()
    changelog = pd.DataFrame(
        [
            {
                "Version": "v0.1",
                "Date": str(date.today()),
                "Input_Workbook": str(INPUT_XLSX.relative_to(ROOT)),
                "Input_Bib": str(INPUT_BIB.relative_to(ROOT)),
                "Input_Workbook_SHA256": sha256_text(INPUT_XLSX.read_bytes().hex()),
                "Input_Bib_SHA256": sha256_text(INPUT_BIB.read_text(encoding="utf-8", errors="replace")),
                "Major_Changes": "Added source/evidence/curation/ML/counterexample workflow sheets; original sheets preserved.",
            }
        ]
    )

    tables = {
        "README": readme,
        "Change_Log": changelog,
        "Data_Dictionary": dictionary,
        "Source_Registry": bib_df,
        "Evidence_Log": evidence,
        "Molecule_Curated": curated,
        "ML_Dataset_View": ml,
        "High_MA_Abiotic": high_abiotic,
        "Low_MA_Bio": low_bio,
        "Curation_Checks": checks,
    }
    for name, df in original.items():
        tables[f"Legacy_{name[:20]}"] = df
    return tables


def reason_excluded(row: pd.Series) -> str:
    reasons = []
    if not str(row["Graph_Level"]).startswith("G3"):
        reasons.append("no_audited_2D_graph")
    if row["Label_Strength"] not in {"strong", "candidate_strong"}:
        reasons.append("weak_or_unchecked_label")
    if pd.isna(row["MA"]):
        reasons.append("MA_missing")
    if row["Source_Validation_Status"] != "source_verified":
        reasons.append("source_not_verified")
    return "; ".join(reasons)


def make_checks(all_df: pd.DataFrame, curated: pd.DataFrame, bib_df: pd.DataFrame, high: pd.DataFrame, low: pd.DataFrame) -> pd.DataFrame:
    rows = [
        ("Legacy All rows", len(all_df)),
        ("BibTeX source entries parsed", len(bib_df)),
        ("Curated rows", len(curated)),
        ("MA missing", int(curated["MA"].isna().sum())),
        ("Canonical_SMILES missing", int(curated["Canonical_SMILES"].isna().sum())),
        ("InChIKey missing", int(curated["InChIKey"].isna().sum())),
        ("Rows needing source link", int((curated["Source_Validation_Status"] == "needs_source_registry_link").sum())),
        ("Rows with possible duplicate or multi-source InChIKey", int((curated["Conflict_Flag"] == "possible_duplicate_or_multi_source").sum())),
        ("High-MA abiotic candidates (MA >= 15)", len(high)),
        ("Low-MA biological candidates (MA < 15)", len(low)),
    ]
    for cls, n in curated["Origin_Class"].value_counts(dropna=False).items():
        rows.append((f"Origin_Class: {cls}", int(n)))
    return pd.DataFrame(rows, columns=["Check", "Value"])


def make_dictionary() -> pd.DataFrame:
    rows = [
        ("Source_Registry", "Source_ID", "Stable ID for a literature/database/source record."),
        ("Source_Registry", "Screening_Status", "candidate / included / excluded; manual review required."),
        ("Evidence_Log", "Record_ID", "One extracted evidence record from a source or legacy sheet."),
        ("Evidence_Log", "Evidence_Location", "Table, figure, supplement, page, or database export location."),
        ("Molecule_Curated", "Origin_Class", "Bio_Strong_Candidate / Abiotic_Strong / Bio_Ambiguous / etc."),
        ("Molecule_Curated", "Origin_Label", "biological / abiotic / ambiguous."),
        ("Molecule_Curated", "Label_Strength", "strong / candidate_strong / weak / unchecked."),
        ("Molecule_Curated", "Nonbiotic_Report_Status", "For biological candidates: none_found / reported / uncertain / not_checked."),
        ("Molecule_Curated", "Graph_Level", "G0 no graph; G1 formula only; G2 candidate; G3 2D graph; G4 stereochemical graph."),
        ("Molecule_Curated", "MA_Use", "MA_primary / MA_sensitivity_or_requires_mol_audit / exclude."),
        ("Molecule_Curated", "Confidence", "A/B/C/D-like curation confidence; current values are migration estimates."),
        ("ML_Dataset_View", "Train_Eligible", "Whether row may enter first-pass ML training after source verification."),
        ("ML_Dataset_View", "Leakage_Group", "Initial anti-leakage grouping from InChIKey first block; replace with scaffold later."),
    ]
    return pd.DataFrame(rows, columns=["Sheet", "Field", "Meaning"])


def make_readme() -> pd.DataFrame:
    rows = [
        ("Purpose", "Research-grade database scaffold for testing MA limitations and preparing future ML biosignature work."),
        ("Principle 1", "Do not treat PubChem presence as a requirement; require an auditable molecular graph or mol/sdf provenance."),
        ("Principle 2", "Separate evidence records from curated molecular identities; repeated molecules across sources are evidence, not noise."),
        ("Principle 3", "Strong claims require source links, evidence locations, and explicit nonbiotic-report searches."),
        ("Current limitation", "Most Source_ID and Evidence_Location fields are blank because legacy rows were not linked to exact papers/tables yet."),
        ("Next step", "Curate 10-20 pilot molecules: link Source_ID, fill Evidence_Location, verify mol files, and update confidence."),
    ]
    return pd.DataFrame(rows, columns=["Item", "Description"])


def write_workbook(tables: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        for sheet_name, df in tables.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

    wb = load_workbook(OUTPUT_XLSX)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        max_col = min(ws.max_column, 40)
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter][: min(ws.max_row, 200)]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 80))
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 45))
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(OUTPUT_XLSX)


if __name__ == "__main__":
    tables = build_tables()
    write_workbook(tables)
    print(OUTPUT_XLSX)
