from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "数据库"
INPUT = DATA_DIR / "5.20完善数据集_research_db_v0.6_source_locations.xlsx"
SOURCE_QUEUE = DATA_DIR / "source_expansion_candidates_v0.1.tsv"
OUTPUT = DATA_DIR / "5.20完善数据集_research_db_v0.7_expansion_planning.xlsx"


SAMPLING_TARGETS = [
    {
        "Sampling_Block": "returned_sample_meteorite_abiotic",
        "Target_Min": 20,
        "Target_Max": 30,
        "Candidate_Source_IDs": "EXP-008; EXP-010; EXP-011",
        "Primary_Purpose": "Strong nonbiological occurrence evidence for biomolecule-like small molecules.",
        "Default_Status": "holdout_until_source_position_and_structure_audit",
        "Key_Risk": "Meteorite data require contamination and analytical-confidence checks.",
    },
    {
        "Sampling_Block": "prebiotic_hydrothermal_ftt_experiment",
        "Target_Min": 15,
        "Target_Max": 25,
        "Candidate_Source_IDs": "EXP-016; EXP-017; EXP-018",
        "Primary_Purpose": "Independent experimental abiotic formation pathways.",
        "Default_Status": "holdout_until_reaction_context_verified",
        "Key_Risk": "Some reported species are condition-level or mixture-level rather than isolated structures.",
    },
    {
        "Sampling_Block": "pah_astrochemical_combustion_complexity",
        "Target_Min": 15,
        "Target_Max": 25,
        "Candidate_Source_IDs": "EXP-012; EXP-013; EXP-014; EXP-015",
        "Primary_Purpose": "High-complexity nonbiological structural pressure tests for MA.",
        "Default_Status": "holdout_until_origin_evidence_separated_from_structure_reference",
        "Key_Risk": "Spectral/theoretical databases do not by themselves prove natural occurrence.",
    },
    {
        "Sampling_Block": "microbial_natural_products",
        "Target_Min": 20,
        "Target_Max": 30,
        "Candidate_Source_IDs": "EXP-001; EXP-002; EXP-003; EXP-004",
        "Primary_Purpose": "Biological high-complexity and medium-complexity structural space.",
        "Default_Status": "holdout_until_species_reference_and_nonbiotic_search",
        "Key_Risk": "Natural-product occurrence is not the same as life-exclusive status.",
    },
    {
        "Sampling_Block": "organism_specific_metabolites",
        "Target_Min": 10,
        "Target_Max": 10,
        "Candidate_Source_IDs": "EXP-005; EXP-006; EXP-007",
        "Primary_Purpose": "Low-MA biological occurrence controls and functional biochemical diversity.",
        "Default_Status": "holdout_until_targeted_nonbiotic_search",
        "Key_Risk": "Small metabolites often have abiotic or environmental reports.",
    },
]


INTAKE_COLUMNS = [
    "Expansion_Record_ID",
    "Expansion_Batch",
    "Candidate_Source_ID",
    "Source_ID",
    "Source_Record_ID",
    "Source_Record_URL",
    "Exact_Source_Position",
    "Molecule_Name",
    "Synonyms",
    "Reported_Formula",
    "Canonical_SMILES",
    "Isomeric_SMILES",
    "InChIKey",
    "CID",
    "Mol_File",
    "Mol_File_SHA256",
    "Structure_Provenance",
    "Source_Domain_Expanded",
    "Reported_Context",
    "Proposed_Origin_Label",
    "Proposed_Label_Strength",
    "Evidence_Tier",
    "Biological_Exclusivity_Status",
    "Nonbiotic_Search_Query",
    "Nonbiotic_Search_Result",
    "MA",
    "MA_Status",
    "MA_Implementation",
    "MA_Code_Commit",
    "MA_Input_File",
    "MA_Input_SHA256",
    "Graph_Level",
    "Graph_Ambiguity",
    "Leakage_Group",
    "ML_Holdout_Reason",
    "Import_Decision",
    "Reviewer_Note",
]


SEED_CANDIDATES = [
    ("SEED-001", "EXP-001; EXP-002", "Microcystin-LR", "biological", "Natural_Product", "High-complexity cyanobacterial natural product; recompute removed placeholder MA."),
    ("SEED-002", "EXP-001; EXP-002", "Nodularin R", "biological", "Natural_Product", "High-complexity cyanobacterial natural product; recompute removed placeholder MA."),
    ("SEED-003", "EXP-002", "Enterobactin", "biological", "Peptide", "Siderophore candidate already useful for low/high MA biological contrast; needs stronger source record."),
    ("SEED-004", "EXP-002; EXP-003", "Desferrioxamine B", "biological", "Natural_Product", "Siderophore-like microbial metabolite with clear structure and biological context to verify."),
    ("SEED-005", "EXP-001; EXP-002", "Vancomycin", "biological", "Natural_Product", "Large glycopeptide natural product for high-complexity biological side."),
    ("SEED-006", "EXP-001; EXP-002", "Erythromycin A", "biological", "Natural_Product", "Macrolide natural product for biological structural diversity."),
    ("SEED-007", "EXP-001; EXP-002", "Rapamycin", "biological", "Natural_Product", "Large microbial natural product; good high-complexity biological candidate."),
    ("SEED-008", "EXP-001; EXP-002", "Amphotericin B", "biological", "Natural_Product", "Polyene macrolide natural product; large structure stress test."),
    ("SEED-009", "EXP-001; EXP-002", "Cyclosporin A", "biological", "Peptide", "Cyclic peptide natural product; tests graph-level and MA computation workflow."),
    ("SEED-010", "EXP-001; EXP-002", "Tetracycline", "biological", "Natural_Product", "Medium-complexity natural product with strong literature trail."),
    ("SEED-011", "EXP-010; EXP-016", "Glycine", "abiotic", "Meteorite_or_Prebiotic", "Same molecule can serve as multiple independent abiotic evidence records."),
    ("SEED-012", "EXP-010; EXP-016", "beta-Alanine", "abiotic", "Meteorite_or_Prebiotic", "Biomolecule-like amino acid with nonbiological occurrence contexts."),
    ("SEED-013", "EXP-010; EXP-016", "Adenine", "abiotic", "Meteorite_or_Prebiotic", "Nucleobase false-positive candidate; exact source and contamination evidence required."),
    ("SEED-014", "EXP-008; EXP-010", "Uracil", "abiotic", "Returned_Sample_or_Meteorite", "Nucleobase candidate; useful across Bennu/Ryugu/meteorite evidence comparison."),
    ("SEED-015", "EXP-010", "Ribose", "abiotic", "Meteorite", "Sugar candidate; import only after exact primary table and analytical confidence check."),
    ("SEED-016", "EXP-010", "Hexamethylenetetramine", "abiotic", "Meteorite", "Complex abiotic organic candidate; useful for nonbiological complexity tests."),
    ("SEED-017", "EXP-012; EXP-013", "Coronene", "abiotic_or_structure_reference", "PAH", "PAH structural pressure test; separate spectral reference from occurrence evidence."),
    ("SEED-018", "EXP-012; EXP-013", "Ovalene", "abiotic_or_structure_reference", "PAH", "Large PAH candidate whose old placeholder MA was removed."),
    ("SEED-019", "EXP-012; EXP-013", "Circumcoronene", "abiotic_or_structure_reference", "PAH", "Very large PAH for MA stress testing; structure provenance must be exact."),
    ("SEED-020", "EXP-012; EXP-013", "Pyrene", "abiotic_or_structure_reference", "PAH", "Smaller PAH baseline and overlap check against existing combustion data."),
]


def main() -> None:
    sheets = pd.read_excel(INPUT, sheet_name=None)
    source_queue = build_source_queue()
    sampling_targets = pd.DataFrame(SAMPLING_TARGETS)
    intake_template = build_intake_template()
    seed_20 = build_seed_candidates(sheets)
    qa_checks = build_qa_checks(source_queue, sampling_targets, seed_20)

    sheets["Expansion_Source_Queue"] = source_queue
    sheets["Expansion_Sampling_Targets"] = sampling_targets
    sheets["Expansion_Intake_Template"] = intake_template
    sheets["Expansion_Seed_20"] = seed_20
    sheets["Expansion_QA_Checks"] = qa_checks
    sheets["Data_Dictionary"] = update_dictionary(sheets["Data_Dictionary"])
    sheets["Curation_Checks"] = update_checks(sheets["Curation_Checks"], source_queue, sampling_targets, seed_20)
    sheets["Change_Log"] = update_changelog(sheets["Change_Log"])
    sheets["README"] = update_internal_readme(sheets["README"])

    write_workbook(sheets)
    print(OUTPUT)


def build_source_queue() -> pd.DataFrame:
    queue = pd.read_csv(SOURCE_QUEUE, sep="\t").fillna("")
    targets = queue["Proposed_Records"].map(parse_target_range)
    queue["Target_Min"] = [t[0] for t in targets]
    queue["Target_Max"] = [t[1] for t in targets]
    queue["Expansion_Batch"] = queue["Priority"].map(lambda p: f"expansion_{str(p).lower()}_pilot")
    queue["Queue_Status"] = queue.apply(queue_status, axis=1)
    queue["Next_Action"] = queue.apply(next_action, axis=1)
    queue["Default_Import_Decision"] = "holdout_pending_audit"
    queue["Retrieved_Date"] = ""
    queue["Screening_Date"] = ""
    queue["Owner"] = ""
    queue["Notes"] = ""
    return queue


def parse_target_range(value: object) -> tuple[int, int]:
    text = str(value).strip()
    numbers = [int(n) for n in re.findall(r"\d+", text)]
    if not numbers:
        return 0, 0
    if len(numbers) == 1:
        return numbers[0], numbers[0]
    return numbers[0], numbers[1]


def queue_status(row: pd.Series) -> str:
    if row.get("Evidence_Tier") == "D":
        return "source_pointer_only_do_not_import_records"
    if row.get("Priority") == "P0":
        return "ready_for_first_screening"
    if row.get("Priority") == "P1":
        return "queue_after_p0_screening"
    return "later_queue"


def next_action(row: pd.Series) -> str:
    source_type = row.get("Source_Type", "")
    tier = row.get("Evidence_Tier", "")
    if tier == "D":
        return "Use only to identify primary sources; do not import molecule records."
    if source_type == "curated_database":
        return "Download or inspect record-level data; sample records with structures and source references."
    if source_type == "primary_literature":
        return "Locate supplementary tables; map molecule names to exact table/page/row positions."
    return "Screen source and decide whether it can support molecule-level evidence."


def build_intake_template() -> pd.DataFrame:
    examples = [
        {
            "Expansion_Record_ID": "EXPAND-00001",
            "Expansion_Batch": "expansion_p0_pilot",
            "Candidate_Source_ID": "EXP-001",
            "Molecule_Name": "example_name_replace_before_import",
            "Proposed_Origin_Label": "biological_or_abiotic",
            "Evidence_Tier": "A/B/C",
            "MA_Status": "pending_or_unverified",
            "ML_Holdout_Reason": "expansion_pilot_holdout_until_full_audit",
            "Import_Decision": "template_do_not_import",
        }
    ]
    df = pd.DataFrame(examples)
    for col in INTAKE_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[INTAKE_COLUMNS]


def build_seed_candidates(sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    existing_names = existing_name_set(sheets["Molecule_Curated"])
    rows = []
    for seed_id, source_ids, name, origin, domain, reason in SEED_CANDIDATES:
        status = "already_present_add_evidence_record" if normalize_name(name) in existing_names else "new_candidate_after_source_audit"
        rows.append(
            {
                "Seed_ID": seed_id,
                "Candidate_Source_IDs": source_ids,
                "Candidate_Name_or_Group": name,
                "Expected_Origin_Label": origin,
                "Proposed_Source_Domain": domain,
                "Current_DB_Status": status,
                "Reason_For_Selection": reason,
                "Required_Verification": required_verification(origin, domain),
                "Suggested_Action": "Create intake row only after exact source position and structure provenance are available.",
                "Import_Status": "not_imported_seed_only",
            }
        )
    return pd.DataFrame(rows)


def existing_name_set(curated: pd.DataFrame) -> set[str]:
    names: set[str] = set()
    for col in ["Preferred_Name", "Original_Name"]:
        if col not in curated.columns:
            continue
        for value in curated[col].dropna():
            names.add(normalize_name(value))
    return names


def normalize_name(value: object) -> str:
    text = str(value).strip().lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def required_verification(origin: str, domain: str) -> str:
    if origin == "biological":
        return "Verify database record, organism/source reference, structure, and targeted nonbiotic-origin search."
    if domain == "PAH":
        return "Verify structure source separately from natural occurrence or combustion/astrochemical evidence."
    return "Verify exact primary table/page/row, analytical confidence, structure, and contamination/context notes."


def build_qa_checks(source_queue: pd.DataFrame, sampling_targets: pd.DataFrame, seed_20: pd.DataFrame) -> pd.DataFrame:
    p0 = int(source_queue["Priority"].eq("P0").sum())
    tier_a = int(source_queue["Evidence_Tier"].eq("A").sum())
    pointer_only = int(source_queue["Evidence_Tier"].eq("D").sum())
    target_min = int(sampling_targets["Target_Min"].sum())
    target_max = int(sampling_targets["Target_Max"].sum())
    seed_present = int(seed_20["Current_DB_Status"].eq("already_present_add_evidence_record").sum())
    seed_new = int(seed_20["Current_DB_Status"].eq("new_candidate_after_source_audit").sum())
    return pd.DataFrame(
        [
            {"Check": "Expansion candidate source rows", "Value": len(source_queue), "Interpretation": "Source groups queued for screening."},
            {"Check": "P0 source rows", "Value": p0, "Interpretation": "First screening sources."},
            {"Check": "Evidence Tier A source rows", "Value": tier_a, "Interpretation": "Likely strongest source groups."},
            {"Check": "Pointer-only source rows", "Value": pointer_only, "Interpretation": "Do not import molecule records directly."},
            {"Check": "Sampling target minimum records", "Value": target_min, "Interpretation": "Lower bound for expansion pilot."},
            {"Check": "Sampling target maximum records", "Value": target_max, "Interpretation": "Upper bound for expansion pilot."},
            {"Check": "Seed candidates already present", "Value": seed_present, "Interpretation": "Add new evidence records rather than duplicate molecules."},
            {"Check": "Seed candidates new or unmatched", "Value": seed_new, "Interpretation": "Potential new molecule entries after audit."},
        ]
    )


def update_dictionary(dictionary: pd.DataFrame) -> pd.DataFrame:
    additions = [
        ("Expansion_Source_Queue", "Candidate_ID", "Identifier for a source group queued for database expansion screening."),
        ("Expansion_Source_Queue", "Queue_Status", "Current status of source-level screening."),
        ("Expansion_Sampling_Targets", "Sampling_Block", "Planned expansion category used to balance evidence domains."),
        ("Expansion_Intake_Template", "Expansion_Record_ID", "Template identifier for a future molecule-level expansion intake row."),
        ("Expansion_Seed_20", "Current_DB_Status", "Whether the seed molecule appears to already exist in the current curated database."),
        ("Expansion_QA_Checks", "Value", "Summary count used to validate the expansion planning workbook."),
    ]
    add_df = pd.DataFrame(additions, columns=["Sheet", "Field", "Meaning"])
    merged = pd.concat([dictionary, add_df], ignore_index=True)
    return merged.drop_duplicates(subset=["Sheet", "Field"], keep="last")


def update_checks(
    checks: pd.DataFrame,
    source_queue: pd.DataFrame,
    sampling_targets: pd.DataFrame,
    seed_20: pd.DataFrame,
) -> pd.DataFrame:
    additions = pd.DataFrame(
        [
            {"Check": "v0.7 expansion source queue rows", "Value": len(source_queue)},
            {"Check": "v0.7 expansion sampling blocks", "Value": len(sampling_targets)},
            {"Check": "v0.7 expansion seed candidates", "Value": len(seed_20)},
            {"Check": "v0.7 expansion target min records", "Value": int(sampling_targets["Target_Min"].sum())},
            {"Check": "v0.7 expansion target max records", "Value": int(sampling_targets["Target_Max"].sum())},
        ]
    )
    return pd.concat([checks, additions], ignore_index=True)


def update_changelog(changelog: pd.DataFrame) -> pd.DataFrame:
    row = {
        "Version": "v0.7_expansion_planning",
        "Date": "2026-06-11",
        "Input_Workbook": str(INPUT.relative_to(ROOT)),
        "Input_Bib": "数据库/Assembly Method.bib",
        "Input_Workbook_SHA256": "",
        "Input_Bib_SHA256": "",
        "Major_Changes": "Added expansion planning sheets: source queue, sampling targets, intake template, 20 seed candidates, and expansion QA checks.",
    }
    return pd.concat([changelog, pd.DataFrame([row])], ignore_index=True)


def update_internal_readme(readme: pd.DataFrame) -> pd.DataFrame:
    rows = [
        {"Section": "Current Version", "Note": "v0.7_expansion_planning adds expansion planning sheets while preserving v0.6 curated evidence records."},
        {"Section": "Expansion", "Note": "Use Expansion_Source_Queue and Expansion_Intake_Template before importing any new molecule records."},
        {"Section": "Expansion", "Note": "Expansion_Seed_20 contains seed candidates only; these rows are not imported evidence."},
    ]
    extra = pd.DataFrame(rows)
    if set(extra.columns).issubset(set(readme.columns)):
        return pd.concat([readme, extra], ignore_index=True)
    return readme


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    preferred_order = [
        "README",
        "Change_Log",
        "Data_Dictionary",
        "Curation_Priority",
        "Expansion_Source_Queue",
        "Expansion_Sampling_Targets",
        "Expansion_Intake_Template",
        "Expansion_Seed_20",
        "Expansion_QA_Checks",
        "Source_Position_Update_Log",
        "Primary_Literature_Audit_Queue",
        "MA_Recompute_Queue",
        "Structure_Audit_Log",
        "Nonbiotic_Search_Log",
        "Source_Link_Task_Queue",
        "MA_Removal_Log",
        "Pilot_Curation_20",
        "Source_Registry",
        "Source_Aliases",
        "Evidence_Log",
        "Molecule_Curated",
        "ML_Dataset_View",
        "High_MA_Abiotic",
        "Low_MA_Bio",
        "Curation_Checks",
    ]
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for name in preferred_order:
            if name in sheets:
                sheets[name].to_excel(writer, sheet_name=name[:31], index=False)
        for name, df in sheets.items():
            if name not in preferred_order:
                df.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(OUTPUT)
    fills = {
        "Expansion_Source_Queue": "4472C4",
        "Expansion_Sampling_Targets": "70AD47",
        "Expansion_Intake_Template": "A5A5A5",
        "Expansion_Seed_20": "ED7D31",
        "Expansion_QA_Checks": "8064A2",
        "Source_Position_Update_Log": "5B9BD5",
        "Primary_Literature_Audit_Queue": "C55A11",
        "Source_Link_Task_Queue": "1F4E78",
    }
    default_fill = "1F4E78"
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = PatternFill("solid", fgColor=fills.get(ws.title, default_fill))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, min(ws.max_column, 80) + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter][: min(ws.max_row, 160)]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 100))
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 60))
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
