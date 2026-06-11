from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "数据库"
INPUT = DATA_DIR / "5.20完善数据集_research_db_v0.1.xlsx"
OUTPUT = DATA_DIR / "5.20完善数据集_research_db_v0.2_pilot20.xlsx"


PILOT = [
    {
        "Curated_ID": "MOL-00001",
        "Pilot_Group": "Bennu direct-return abiotic low-MA",
        "Source_ID": "SRC-0013",
        "Evidence_Location": "Glavin et al. 2025, Fig. 3 / amino-acid LC-FD/HRMS; Nature page lines 171-177 identify glycine chromatogram.",
        "Evidence_Level": "direct_detection_compound_level",
        "Pilot_Rationale": "Abiotic returned asteroid sample; low MA glycine prevents interpreting low MA as absence of prebiotic relevance.",
        "Source_Note": "Bennu samples were returned by OSIRIS-REx and analysed under controlled conditions; paper reports amino acids and racemic/non-protein amino-acid context.",
        "Curation_Decision": "keep_for_MA_and_ML_as_abiotic_strong_after_source_check",
        "Confidence": "A_minus_source_linked_group_compound",
    },
    {
        "Curated_ID": "MOL-00004",
        "Pilot_Group": "Bennu direct-return abiotic low-MA",
        "Source_ID": "SRC-0013",
        "Evidence_Location": "Glavin et al. 2025, Fig. 3 / amino-acid LC-FD/HRMS; Nature page lines 171-177 identify beta-alanine chromatogram.",
        "Evidence_Level": "direct_detection_compound_level",
        "Pilot_Rationale": "Abiotic returned asteroid sample with low MA amino-acid isomer.",
        "Source_Note": "Use exact supplementary table during final audit for abundance and enantiomeric ratio.",
        "Curation_Decision": "keep_for_MA_and_ML_as_abiotic_strong_after_source_check",
        "Confidence": "A_minus_source_linked_group_compound",
    },
    {
        "Curated_ID": "MOL-00045",
        "Pilot_Group": "Bennu direct-return abiotic nucleobase",
        "Source_ID": "SRC-0013",
        "Evidence_Location": "Glavin et al. 2025 abstract reports N-heterocycles including all five DNA/RNA nucleobases.",
        "Evidence_Level": "direct_detection_group_level",
        "Pilot_Rationale": "Abiotic returned asteroid sample containing a biological nucleobase class member at low MA.",
        "Source_Note": "Compound-level abundance should be checked against supplementary data before publication.",
        "Curation_Decision": "keep_for_MA; ML label strong only after compound-level supplement audit",
        "Confidence": "B_plus_group_source_linked",
    },
    {
        "Curated_ID": "MOL-00048",
        "Pilot_Group": "Bennu direct-return abiotic nucleobase",
        "Source_ID": "SRC-0013",
        "Evidence_Location": "Glavin et al. 2025 abstract reports N-heterocycles including all five DNA/RNA nucleobases.",
        "Evidence_Level": "direct_detection_group_level",
        "Pilot_Rationale": "Abiotic occurrence of adenine-like biological molecule challenges any life-only interpretation of biomolecular identity.",
        "Source_Note": "Needs supplementary-table confirmation for adenine row and possible blank correction.",
        "Curation_Decision": "keep_for_MA; ML label strong only after compound-level supplement audit",
        "Confidence": "B_plus_group_source_linked",
    },
    {
        "Curated_ID": "MOL-00078",
        "Pilot_Group": "Bennu direct-return abiotic PAH",
        "Source_ID": "SRC-0013",
        "Evidence_Location": "Glavin et al. 2025 abstract reports PAHs; Results section reports PAHs/alkylated PAHs in methanol extracts.",
        "Evidence_Level": "direct_detection_group_level",
        "Pilot_Rationale": "Abiotic PAH occurrence provides low-to-mid MA nonbiological background chemistry.",
        "Source_Note": "Naphthalene specifically requires supplementary or source-data confirmation.",
        "Curation_Decision": "keep_as_pilot_but_mark_compound_level_supplement_needed",
        "Confidence": "B_group_source_linked",
    },
    {
        "Curated_ID": "MOL-00084",
        "Pilot_Group": "Ryugu direct-return abiotic low-MA",
        "Source_ID": "SRC-0048",
        "Evidence_Location": "Naraoka et al. 2023, Science, DOI 10.1126/science.abn9033; soluble organic molecules in Ryugu samples.",
        "Evidence_Level": "direct_detection_literature_linked_needs_table",
        "Pilot_Rationale": "Independent asteroid-return sample provides external validation of low-MA abiotic amino acids.",
        "Source_Note": "Science page inaccessible to automated fetch; use paper/supplement locally for exact table.",
        "Curation_Decision": "keep_for_MA; verify exact table before final analysis",
        "Confidence": "B_source_linked_table_needed",
    },
    {
        "Curated_ID": "MOL-00086",
        "Pilot_Group": "Ryugu direct-return abiotic low-MA",
        "Source_ID": "SRC-0048",
        "Evidence_Location": "Naraoka et al. 2023, Science, DOI 10.1126/science.abn9033; soluble organic molecules in Ryugu samples.",
        "Evidence_Level": "direct_detection_literature_linked_needs_table",
        "Pilot_Rationale": "Low-MA abiotic beta-alanine supports the need to separate prebiotic relevance from life specificity.",
        "Source_Note": "Verify exact enantiomer/isomer reporting from supplementary table.",
        "Curation_Decision": "keep_for_MA; verify exact table before final analysis",
        "Confidence": "B_source_linked_table_needed",
    },
    {
        "Curated_ID": "MOL-00116",
        "Pilot_Group": "Ryugu direct-return abiotic nucleobase",
        "Source_ID": "SRC-0048",
        "Evidence_Location": "Naraoka et al. 2023 source set; Ryugu organic molecule literature reports uracil/vitamin B3 in returned samples.",
        "Evidence_Level": "direct_detection_literature_linked_needs_specific_source",
        "Pilot_Rationale": "Abiotic occurrence of uracil is important for preventing biological-name leakage in ML labels.",
        "Source_Note": "Add the dedicated Oba et al. 2023 uracil source in the next source-registry pass.",
        "Curation_Decision": "keep_as_pilot; source registry should be expanded",
        "Confidence": "B_minus_specific_source_needed",
    },
    {
        "Curated_ID": "MOL-00117",
        "Pilot_Group": "Ryugu direct-return abiotic vitamin precursor",
        "Source_ID": "SRC-0048",
        "Evidence_Location": "Naraoka et al. 2023 source set; Ryugu organic molecule literature reports vitamin B3/nicotinic acid in returned samples.",
        "Evidence_Level": "direct_detection_literature_linked_needs_specific_source",
        "Pilot_Rationale": "Biologically familiar molecule in abiotic asteroid context; useful for ML leakage controls.",
        "Source_Note": "Add the dedicated Ryugu uracil/vitamin B3 source in the next source-registry pass.",
        "Curation_Decision": "keep_as_pilot; source registry should be expanded",
        "Confidence": "B_minus_specific_source_needed",
    },
    {
        "Curated_ID": "MOL-00310",
        "Pilot_Group": "High-MA abiotic methane-pyrolysis PAH",
        "Source_ID": "SRC-0031",
        "Evidence_Location": "Khrabry et al. 2024 / arXiv 2311.00910; abstract states methane pyrolysis forms large PAHs and mechanism includes HACA pathways up to 37 aromatic rings.",
        "Evidence_Level": "abiotic_process_model_group_level",
        "Pilot_Rationale": "MA >= 15 abiotic PAH counterexample candidate.",
        "Source_Note": "Structure is clear; final publication should map exact species name to mechanism species label.",
        "Curation_Decision": "keep_as_high_MA_abiotic_counterexample_candidate",
        "Confidence": "B_group_process_source_linked",
    },
    {
        "Curated_ID": "MOL-00311",
        "Pilot_Group": "High-MA abiotic methane-pyrolysis PAH",
        "Source_ID": "SRC-0031",
        "Evidence_Location": "Khrabry et al. 2024 / arXiv 2311.00910; abstract states methane pyrolysis forms abundant large PAHs under long residence times.",
        "Evidence_Level": "abiotic_process_model_group_level",
        "Pilot_Rationale": "MA = 18 abiotic PAH counterexample candidate.",
        "Source_Note": "Needs exact species-label crosswalk before final table.",
        "Curation_Decision": "keep_as_high_MA_abiotic_counterexample_candidate",
        "Confidence": "B_group_process_source_linked",
    },
    {
        "Curated_ID": "MOL-00312",
        "Pilot_Group": "High-MA abiotic methane-pyrolysis PAH",
        "Source_ID": "SRC-0031",
        "Evidence_Location": "Khrabry et al. 2024 / arXiv 2311.00910; model designed for carbon nanostructure synthesis and PAH growth.",
        "Evidence_Level": "abiotic_process_model_group_level",
        "Pilot_Rationale": "MA = 19 abiotic PAH counterexample candidate.",
        "Source_Note": "Exact mechanism mapping still required.",
        "Curation_Decision": "keep_as_high_MA_abiotic_counterexample_candidate",
        "Confidence": "B_group_process_source_linked",
    },
    {
        "Curated_ID": "MOL-00313",
        "Pilot_Group": "High-MA abiotic methane-pyrolysis PAH",
        "Source_ID": "SRC-0031",
        "Evidence_Location": "Khrabry et al. 2024 / arXiv 2311.00910; mechanism includes large PAHs up to 37 aromatic rings.",
        "Evidence_Level": "abiotic_process_model_group_level",
        "Pilot_Rationale": "MA = 20 abiotic PAH counterexample candidate.",
        "Source_Note": "Excellent stress-test molecule for MA threshold, but species crosswalk is required.",
        "Curation_Decision": "keep_as_high_MA_abiotic_counterexample_candidate",
        "Confidence": "B_group_process_source_linked",
    },
    {
        "Curated_ID": "MOL-00212",
        "Pilot_Group": "Low-MA biomarker candidate",
        "Source_ID": "SRC-0053",
        "Evidence_Location": "Peters et al. 2007 Biomarkers and Isotopes in the Environment and Human History / biomarker reference source.",
        "Evidence_Level": "biomarker_reference_class_level",
        "Pilot_Rationale": "Low MA molecule often used in organic geochemistry; useful as low-MA biological-origin candidate.",
        "Source_Note": "Pristane can derive from biological precursors through diagenetic alteration; do not claim life-exclusive without targeted search.",
        "Curation_Decision": "keep_as_low_MA_bio_candidate_not_life_exclusive",
        "Confidence": "B_candidate_requires_nonbiotic_search",
    },
    {
        "Curated_ID": "MOL-00213",
        "Pilot_Group": "Low-MA biomarker candidate",
        "Source_ID": "SRC-0053",
        "Evidence_Location": "Peters et al. 2007 Biomarkers and Isotopes in the Environment and Human History / biomarker reference source.",
        "Evidence_Level": "biomarker_reference_class_level",
        "Pilot_Rationale": "Low MA phytane challenges simple low-MA exclusion, but source interpretation must be contextual.",
        "Source_Note": "Phytane/pristane are not clean life-exclusive labels; record redox/source-rock context during final curation.",
        "Curation_Decision": "keep_as_low_MA_bio_candidate_not_life_exclusive",
        "Confidence": "B_candidate_requires_nonbiotic_search",
    },
    {
        "Curated_ID": "MOL-00214",
        "Pilot_Group": "Low-MA biomarker candidate",
        "Source_ID": "SRC-0053",
        "Evidence_Location": "Peters et al. 2007 Biomarkers and Isotopes in the Environment and Human History / biomarker reference source.",
        "Evidence_Level": "biomarker_reference_class_level",
        "Pilot_Rationale": "Low MA hydrocarbon biomarker candidate.",
        "Source_Note": "Needs targeted nonbiotic occurrence search before use as strong biological training label.",
        "Curation_Decision": "keep_as_low_MA_bio_candidate",
        "Confidence": "B_candidate_requires_nonbiotic_search",
    },
    {
        "Curated_ID": "MOL-00215",
        "Pilot_Group": "Low-MA biomarker candidate",
        "Source_ID": "SRC-0053",
        "Evidence_Location": "Peters et al. 2007 Biomarkers and Isotopes in the Environment and Human History / biomarker reference source.",
        "Evidence_Level": "biomarker_reference_class_level",
        "Pilot_Rationale": "Sterane-related low-MA candidate for biological/geological biomarker discussion.",
        "Source_Note": "Needs source-specific context and nonbiotic search.",
        "Curation_Decision": "keep_as_low_MA_bio_candidate",
        "Confidence": "B_candidate_requires_nonbiotic_search",
    },
    {
        "Curated_ID": "MOL-00216",
        "Pilot_Group": "Low-MA biomarker candidate",
        "Source_ID": "SRC-0053",
        "Evidence_Location": "Peters et al. 2007 Biomarkers and Isotopes in the Environment and Human History / biomarker reference source.",
        "Evidence_Level": "biomarker_reference_class_level",
        "Pilot_Rationale": "Low MA triterpane biomarker candidate.",
        "Source_Note": "Needs source-specific geochemical context and nonbiotic search.",
        "Curation_Decision": "keep_as_low_MA_bio_candidate",
        "Confidence": "B_candidate_requires_nonbiotic_search",
    },
    {
        "Curated_ID": "MOL-00218",
        "Pilot_Group": "Low-MA biomarker candidate",
        "Source_ID": "SRC-0053",
        "Evidence_Location": "Peters et al. 2007 Biomarkers and Isotopes in the Environment and Human History / biomarker reference source.",
        "Evidence_Level": "biomarker_reference_class_level",
        "Pilot_Rationale": "Low MA hopanoid-related biomarker candidate; good counterexample if biological specificity is verified.",
        "Source_Note": "Prioritize targeted search for abiotic reports before final ML label.",
        "Curation_Decision": "keep_as_low_MA_bio_candidate",
        "Confidence": "B_candidate_requires_nonbiotic_search",
    },
    {
        "Curated_ID": "MOL-00240",
        "Pilot_Group": "Low-MA biological siderophore",
        "Source_ID": "SRC-NEW-SIDEROPHORE-REVIEW",
        "Evidence_Location": "Needs primary siderophore source; current row selected because enterobactin is a microbial iron-transport molecule.",
        "Evidence_Level": "biological_class_known_needs_primary_source",
        "Pilot_Rationale": "Biological natural product with MA below 15; useful for low-MA biological counterexample after source registration.",
        "Source_Note": "Add a primary enterobactin/siderophore review to Source_Registry before publication.",
        "Curation_Decision": "keep_as_pilot_but_do_not_train_until_source_added",
        "Confidence": "C_primary_source_missing",
    },
]


def main() -> None:
    sheets = pd.read_excel(INPUT, sheet_name=None)
    pilot = pd.DataFrame(PILOT)

    curated = sheets["Molecule_Curated"].copy()
    evidence = sheets["Evidence_Log"].copy()
    ml = sheets["ML_Dataset_View"].copy()

    for df, cols in [
        (curated, ["Source_Validation_Status", "Confidence", "Curation_Note", "Nonbiotic_Report_Status", "Label_Strength", "MA_Use"]),
        (evidence, ["Source_ID", "Evidence_Location", "Evidence_Confidence", "Notes"]),
    ]:
        for col in cols:
            if col in df.columns:
                df[col] = df[col].astype("object")

    pilot_by_id = pilot.set_index("Curated_ID")
    record_by_curated = curated.set_index("Curated_ID")["Record_IDs"].to_dict()

    for cid, row in pilot_by_id.iterrows():
        cur_mask = curated["Curated_ID"] == cid
        if not cur_mask.any():
            continue
        record_id = record_by_curated[cid]
        ev_mask = evidence["Record_ID"] == record_id

        curated.loc[cur_mask, "Source_Validation_Status"] = "pilot_curated_source_linked"
        curated.loc[cur_mask, "Confidence"] = row["Confidence"]
        curated.loc[cur_mask, "Curation_Note"] = row["Curation_Decision"] + " | " + row["Source_Note"]
        if str(row["Pilot_Group"]).startswith("Low-MA biomarker"):
            curated.loc[cur_mask, "Nonbiotic_Report_Status"] = "not_checked_targeted_search_required"
            curated.loc[cur_mask, "Label_Strength"] = "candidate_strong"
        if "siderophore" in str(row["Pilot_Group"]).lower():
            curated.loc[cur_mask, "Source_Validation_Status"] = "pilot_needs_new_primary_source"
            curated.loc[cur_mask, "Label_Strength"] = "weak_until_source_added"
        if "High-MA abiotic" in row["Pilot_Group"]:
            curated.loc[cur_mask, "MA_Use"] = "MA_primary_counterexample_candidate"

        evidence.loc[ev_mask, "Source_ID"] = row["Source_ID"]
        evidence.loc[ev_mask, "Evidence_Location"] = row["Evidence_Location"]
        evidence.loc[ev_mask, "Evidence_Confidence"] = row["Confidence"]
        evidence.loc[ev_mask, "Notes"] = row["Pilot_Rationale"] + " | " + row["Source_Note"]

    # ML view remains conservative: source-linked rows are still provisional until
    # exact table/supplement locations and leakage groups are finalized.
    ml = ml.drop(columns=[c for c in ["Pilot_Curation_Status", "Pilot_Group"] if c in ml.columns])
    ml = ml.merge(
        pilot[["Curated_ID", "Pilot_Group", "Curation_Decision"]],
        on="Curated_ID",
        how="left",
    )
    ml["Pilot_Curation_Status"] = ml["Pilot_Group"].notna().map({True: "pilot_curated", False: ""})
    ml.loc[ml["Pilot_Curation_Status"].eq("pilot_curated"), "Train_Eligible"] = "pilot_holdout_until_full_source_audit"
    ml.loc[ml["Pilot_Curation_Status"].eq("pilot_curated"), "Reason_Excluded"] = (
        "pilot row; hold out until exact evidence location, mol provenance, and leakage group are audited"
    )

    pilot = pilot.merge(
        curated[
            [
                "Curated_ID",
                "Record_IDs",
                "Preferred_Name",
                "Formula",
                "Canonical_SMILES",
                "InChIKey",
                "CID",
                "Origin",
                "Origin_Label",
                "Origin_Class",
                "Label_Strength",
                "MA",
                "Graph_Level",
                "MA_Use",
                "Nonbiotic_Report_Status",
            ]
        ],
        on="Curated_ID",
        how="left",
    )
    ordered = [
        "Curated_ID",
        "Record_IDs",
        "Preferred_Name",
        "Formula",
        "Origin",
        "Origin_Label",
        "Origin_Class",
        "Label_Strength",
        "MA",
        "Graph_Level",
        "MA_Use",
        "Source_ID",
        "Evidence_Level",
        "Evidence_Location",
        "Pilot_Group",
        "Pilot_Rationale",
        "Curation_Decision",
        "Confidence",
        "Nonbiotic_Report_Status",
        "Source_Note",
        "Canonical_SMILES",
        "InChIKey",
        "CID",
    ]
    pilot = pilot[ordered]

    sheets["Molecule_Curated"] = curated
    sheets["Evidence_Log"] = evidence
    sheets["ML_Dataset_View"] = ml
    sheets["Pilot_Curation_20"] = pilot

    checks = sheets["Curation_Checks"].copy()
    extra = pd.DataFrame(
        [
            {"Check": "Pilot curated rows", "Value": len(pilot)},
            {"Check": "Pilot rows needing primary/source supplement", "Value": int(pilot["Confidence"].astype(str).str.contains("needed|missing|table").sum())},
            {"Check": "Pilot high-MA abiotic rows", "Value": int(pilot["Pilot_Group"].astype(str).str.contains("High-MA abiotic").sum())},
            {"Check": "Pilot low-MA biological/biomarker rows", "Value": int(pilot["Pilot_Group"].astype(str).str.contains("Low-MA bio|siderophore").sum())},
        ]
    )
    sheets["Curation_Checks"] = pd.concat([checks, extra], ignore_index=True)

    changelog = sheets["Change_Log"].copy()
    changelog = pd.concat(
        [
            changelog,
            pd.DataFrame(
                [
                    {
                        "Version": "v0.2_pilot20",
                        "Date": "2026-06-11",
                        "Input_Workbook": str(INPUT.relative_to(ROOT)),
                        "Input_Bib": "数据库/Assembly Method.bib",
                        "Input_Workbook_SHA256": "",
                        "Input_Bib_SHA256": "",
                        "Major_Changes": "Added Pilot_Curation_20 and linked 20 selected molecules to source/evidence notes; ML rows held out until full audit.",
                    }
                ]
            ),
        ],
        ignore_index=True,
    )
    sheets["Change_Log"] = changelog

    write_workbook(sheets)
    print(OUTPUT)


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    preferred_order = [
        "README",
        "Change_Log",
        "Data_Dictionary",
        "Pilot_Curation_20",
        "Source_Registry",
        "Evidence_Log",
        "Molecule_Curated",
        "ML_Dataset_View",
        "High_MA_Abiotic",
        "Low_MA_Bio",
        "Curation_Checks",
    ]
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for name in preferred_order:
            if name in sheets:
                sheets[name].to_excel(writer, sheet_name=name[:31], index=False)
        for name, df in sheets.items():
            if name not in preferred_order:
                df.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(OUTPUT)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    pilot_fill = PatternFill("solid", fgColor="6B8E23")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = pilot_fill if ws.title == "Pilot_Curation_20" else header_fill
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, min(ws.max_column, 50) + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter][: min(ws.max_row, 120)]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 90))
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 55))
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
