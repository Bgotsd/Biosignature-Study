from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "数据库"
SOURCE_XLSX = DATA_DIR / "5.20完善数据集.xlsx"
RESEARCH_XLSX = DATA_DIR / "5.20完善数据集_research_db_v0.3_sources_dedup.xlsx"
CORRECTED_SOURCE = DATA_DIR / "5.20完善数据集_ma_corrected.xlsx"
CORRECTED_RESEARCH = DATA_DIR / "5.20完善数据集_research_db_v0.4_ma_corrected.xlsx"

HIGHLIGHT_RGB = "FFFFFF00"


def cell_is_highlighted(cell) -> bool:
    fill = cell.fill
    if not fill or not fill.fill_type or not fill.fgColor:
        return False
    return str(fill.fgColor.rgb) == HIGHLIGHT_RGB


def highlighted_all_rows() -> list[dict]:
    wb = load_workbook(SOURCE_XLSX, data_only=False)
    ws = wb["All"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    ma_col = headers.index("MA") + 1
    rows = []
    for r in range(2, ws.max_row + 1):
        if any(cell_is_highlighted(ws.cell(r, c)) for c in range(1, ws.max_column + 1)):
            record_id = f"REC-{r - 1:05d}"
            item = {
                "Record_ID": record_id,
                "Original_All_Row": r,
                "Original_Name": ws.cell(r, headers.index("Original_Name") + 1).value,
                "Origin": ws.cell(r, headers.index("Origin") + 1).value,
                "Formula": ws.cell(r, headers.index("Formula") + 1).value,
                "Old_MA": ws.cell(r, ma_col).value,
                "Removal_Reason": "Yellow-highlighted in original All sheet; user indicated MA was randomly filled because exact MA calculation was not feasible locally.",
            }
            rows.append(item)
    return rows


def clear_source_workbook(rows: list[dict]) -> None:
    wb = load_workbook(SOURCE_XLSX)
    keys = {(r["Original_Name"], r["Origin"]) for r in rows}
    names = {r["Original_Name"] for r in rows}
    for ws in wb.worksheets:
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        if "MA" not in headers or "Original_Name" not in headers:
            continue
        ma_col = headers.index("MA") + 1
        name_col = headers.index("Original_Name") + 1
        origin_col = headers.index("Origin") + 1 if "Origin" in headers else None
        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, name_col).value
            origin = ws.cell(r, origin_col).value if origin_col else None
            if (name, origin) in keys or (ws.title != "All" and name in names):
                ws.cell(r, ma_col).value = None
                # Preserve a visual audit marker in corrected source workbook.
                ws.cell(r, ma_col).fill = PatternFill("solid", fgColor="FFF2CC")
    wb.save(CORRECTED_SOURCE)


def clear_research_workbook(rows: list[dict]) -> None:
    sheets = pd.read_excel(RESEARCH_XLSX, sheet_name=None)
    log = pd.DataFrame(rows)
    affected_records = set(log["Record_ID"])
    affected_names = set(log["Original_Name"])

    for sheet_name in ["Molecule_Curated", "ML_Dataset_View", "Pilot_Curation_20", "High_MA_Abiotic", "Low_MA_Bio"]:
        if sheet_name not in sheets or "MA" not in sheets[sheet_name].columns:
            continue
        df = sheets[sheet_name].copy()
        mask = pd.Series(False, index=df.index)
        if "Record_IDs" in df.columns:
            mask |= df["Record_IDs"].isin(affected_records)
        if "Record_ID" in df.columns:
            mask |= df["Record_ID"].isin(affected_records)
        if "Preferred_Name" in df.columns:
            # Use this only as a fallback for derived candidate sheets.
            if "Record_IDs" not in df.columns and "Record_ID" not in df.columns:
                mask |= df["Preferred_Name"].isin(affected_names)
        if "Original_Name" in df.columns and "Record_IDs" not in df.columns and "Record_ID" not in df.columns:
            mask |= df["Original_Name"].isin(affected_names)
        df.loc[mask, "MA"] = pd.NA

        if sheet_name == "Molecule_Curated":
            df.loc[mask, "MA_Use"] = "exclude_until_MA_recomputed"
            df.loc[mask, "Confidence"] = "C_MA_removed_requires_recalculation"
            note = "MA removed in v0.4 because original yellow-highlighted value was randomly filled; recompute with auditable algorithm/server run before analysis."
            df.loc[mask, "Curation_Note"] = df.loc[mask, "Curation_Note"].fillna("").map(lambda x: (str(x) + " | " + note).strip(" |"))
        elif sheet_name == "ML_Dataset_View":
            df.loc[mask, "Train_Eligible"] = "no"
            df.loc[mask, "Reason_Excluded"] = "MA removed: yellow-highlighted original value was randomly filled; recompute before ML use"
        elif sheet_name == "Pilot_Curation_20":
            df.loc[mask, "Curation_Decision"] = df.loc[mask, "Curation_Decision"].fillna("").map(
                lambda x: (str(x) + " | MA removed; no longer valid as MA counterexample until recomputed.").strip(" |")
            )
            df.loc[mask, "Confidence"] = "C_MA_removed_requires_recalculation"
        sheets[sheet_name] = df

    if "Evidence_Log" in sheets:
        ev = sheets["Evidence_Log"].copy()
        mask = ev["Record_ID"].isin(affected_records)
        ev.loc[mask, "Notes"] = ev.loc[mask, "Notes"].fillna("").map(
            lambda x: (str(x) + " | MA removed in v0.4; original value was yellow-highlighted/random.").strip(" |")
        )
        sheets["Evidence_Log"] = ev

    # Recompute candidate views from corrected Molecule_Curated.
    curated = sheets["Molecule_Curated"]
    numeric_ma = pd.to_numeric(curated["MA"], errors="coerce")
    high = curated[(curated["Origin_Label"] == "abiotic") & (numeric_ma >= 15)].copy()
    if not high.empty:
        high["Why_it_challenges_MA_threshold"] = "Abiotic-labeled molecule with MA >= 15; requires source verification before publication."
    sheets["High_MA_Abiotic"] = high

    low = curated[(curated["Origin_Label"] == "biological") & (numeric_ma < 15)].copy()
    if not low.empty:
        low["Why_it_challenges_low_MA_exclusion"] = "Biological-labeled candidate with MA < 15; requires nonbiotic-report search before strong claim."
    sheets["Low_MA_Bio"] = low

    sheets["MA_Removal_Log"] = log

    checks = sheets["Curation_Checks"].copy()
    checks = pd.concat(
        [
            checks,
            pd.DataFrame(
                [
                    {"Check": "v0.4 highlighted/random MA values removed", "Value": len(log)},
                    {"Check": "v0.4 High-MA abiotic candidates after removal", "Value": len(high)},
                    {"Check": "v0.4 Low-MA biological candidates after removal", "Value": len(low)},
                ]
            ),
        ],
        ignore_index=True,
    )
    sheets["Curation_Checks"] = checks

    changelog = sheets["Change_Log"].copy()
    changelog = pd.concat(
        [
            changelog,
            pd.DataFrame(
                [
                    {
                        "Version": "v0.4_ma_corrected",
                        "Date": "2026-06-11",
                        "Input_Workbook": str(RESEARCH_XLSX.relative_to(ROOT)),
                        "Input_Bib": "数据库/Assembly Method.bib",
                        "Input_Workbook_SHA256": "",
                        "Input_Bib_SHA256": "",
                        "Major_Changes": "Removed MA values for 51 yellow-highlighted original rows whose MA values were randomly filled; added MA_Removal_Log and regenerated high/low MA candidate views.",
                    }
                ]
            ),
        ],
        ignore_index=True,
    )
    sheets["Change_Log"] = changelog

    write_workbook(sheets, CORRECTED_RESEARCH)


def write_workbook(sheets: dict[str, pd.DataFrame], output: Path) -> None:
    preferred_order = [
        "README",
        "Change_Log",
        "Data_Dictionary",
        "MA_Removal_Log",
        "Pilot_Curation_20",
        "Source_Registry",
        "Source_Aliases",
        "Evidence_Log",
        "Molecule_Curated",
        "ML_Dataset_View",
        "High_MA_Abiotic",
        "Low_MA_Bio",
        "Curation_Checks",
    ]
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name in preferred_order:
            if name in sheets:
                sheets[name].to_excel(writer, sheet_name=name[:31], index=False)
        for name, df in sheets.items():
            if name not in preferred_order:
                df.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(output)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    removal_fill = PatternFill("solid", fgColor="C00000")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = removal_fill if ws.title == "MA_Removal_Log" else header_fill
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, min(ws.max_column, 55) + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter][: min(ws.max_row, 150)]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 90))
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 55))
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(output)


def main() -> None:
    rows = highlighted_all_rows()
    clear_source_workbook(rows)
    clear_research_workbook(rows)
    print(f"Removed {len(rows)} highlighted/random MA values")
    print(CORRECTED_SOURCE)
    print(CORRECTED_RESEARCH)


if __name__ == "__main__":
    main()
