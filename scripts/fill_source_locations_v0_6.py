from __future__ import annotations

import json
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "数据库"
INPUT = DATA_DIR / "5.20完善数据集_research_db_v0.5_workflow_enhanced.xlsx"
OUTPUT = DATA_DIR / "5.20完善数据集_research_db_v0.6_source_locations.xlsx"


LOCAL_SOURCES = [
    {
        "Source_ID": "SRC-0073",
        "BibTeX_Key": "muchowskaSynthesisPrebioticKrebs2019_local",
        "Source_Type": "article_and_local_reaction_json",
        "Title": "Prebiotic reaction network source extracted from pba condition files",
        "Authors": "",
        "Year": "2019",
        "Journal": "Nature",
        "DOI_or_URL": "10.1038/s41586-019-1151-1",
        "PMID": "",
        "Source_Domain": "Prebiotic",
        "Search_Query": "",
        "Why_Relevant": "Primary DOI recorded in Python-MA/prebio/conditions/pba-*.json source fields.",
        "Access_Date": "",
        "Screening_Status": "included",
        "Exclusion_Reason": "",
        "Notes": "Added during v0.6 source-location pass; verify title/authors against publisher metadata before final citation.",
    },
    {
        "Source_ID": "SRC-LOCAL-COMBUSTION-A",
        "BibTeX_Key": "localCombustionASeries",
        "Source_Type": "local_curated_structure_set",
        "Title": "Local combustion A-series candidate structures",
        "Authors": "",
        "Year": "",
        "Journal": "",
        "DOI_or_URL": "PAH ChemDraw/ and Legacy_Combustion",
        "PMID": "",
        "Source_Domain": "Combustion",
        "Search_Query": "",
        "Why_Relevant": "Local A-series structures/labels used in the legacy workbook; primary literature still pending.",
        "Access_Date": "",
        "Screening_Status": "included_local_pending_primary",
        "Exclusion_Reason": "",
        "Notes": "Do not use as publication-grade source until primary combustion paper/table is linked.",
    },
    {
        "Source_ID": "SRC-LOCAL-PEPTIDE",
        "BibTeX_Key": "localPeptideBiopepList",
        "Source_Type": "local_curated_molecule_list",
        "Title": "Local peptide and siderophore molecule list",
        "Authors": "",
        "Year": "",
        "Journal": "",
        "DOI_or_URL": "Python-MA/Bennu/Peptide.txt; Legacy_Peptide",
        "PMID": "",
        "Source_Domain": "Peptide",
        "Search_Query": "",
        "Why_Relevant": "Local list used to build peptide/siderophore rows; primary biological source still pending.",
        "Access_Date": "",
        "Screening_Status": "included_local_pending_primary",
        "Exclusion_Reason": "",
        "Notes": "Use only for local provenance; add primary peptide/siderophore literature before strong biological claims.",
    },
]


def main() -> None:
    sheets = pd.read_excel(INPUT, sheet_name=None)
    source_registry = add_local_sources(sheets["Source_Registry"].copy())
    evidence = sheets["Evidence_Log"].copy()
    curated = sheets["Molecule_Curated"].copy()

    legacy_row_maps = build_legacy_row_maps(sheets)
    text_line_maps = build_text_line_maps()
    prebio_conditions = build_prebio_condition_index()

    update_log_rows = []
    for idx, row in evidence.iterrows():
        record_id = row["Record_ID"]
        cur = curated[curated["Record_IDs"].eq(record_id)]
        if cur.empty:
            continue
        cur_row = cur.iloc[0]
        origin = str(cur_row.get("Origin", ""))
        name = str(cur_row.get("Original_Name", cur_row.get("Preferred_Name", "")))
        preferred = str(cur_row.get("Preferred_Name", name))
        all_row = record_number(record_id) + 1

        source_id, location, status = infer_source_location(
            origin=origin,
            name=name,
            preferred=preferred,
            cur_row=cur_row,
            all_row=all_row,
            legacy_row_maps=legacy_row_maps,
            text_line_maps=text_line_maps,
            prebio_conditions=prebio_conditions,
        )

        old_source = row.get("Source_ID")
        old_location = row.get("Evidence_Location")

        if blank(old_source):
            evidence.at[idx, "Source_ID"] = source_id
        if blank(old_location):
            evidence.at[idx, "Evidence_Location"] = location
        elif "Legacy_All row" not in str(old_location):
            evidence.at[idx, "Evidence_Location"] = str(old_location) + " | " + location

        evidence.at[idx, "Evidence_Position_Status"] = status
        update_log_rows.append(
            {
                "Record_ID": record_id,
                "Curated_ID": cur_row.get("Curated_ID"),
                "Preferred_Name": preferred,
                "Origin": origin,
                "Assigned_Source_ID": evidence.at[idx, "Source_ID"],
                "Evidence_Position_Status": status,
                "Evidence_Location": evidence.at[idx, "Evidence_Location"],
            }
        )

    sheets["Source_Registry"] = source_registry
    sheets["Evidence_Log"] = evidence
    sheets["Source_Position_Update_Log"] = pd.DataFrame(update_log_rows)
    sheets["Source_Link_Task_Queue"] = rebuild_source_link_queue(evidence, curated)
    sheets["Primary_Literature_Audit_Queue"] = build_primary_literature_audit_queue(evidence, curated)
    sheets["Data_Dictionary"] = update_dictionary(sheets["Data_Dictionary"])
    sheets["Curation_Checks"] = update_checks(sheets["Curation_Checks"], evidence, sheets)
    sheets["Change_Log"] = update_changelog(sheets["Change_Log"])

    write_workbook(sheets)
    print(OUTPUT)


def blank(value: object) -> bool:
    return pd.isna(value) or str(value).strip() == ""


def record_number(record_id: str) -> int:
    return int(str(record_id).split("-")[-1])


def add_local_sources(source_registry: pd.DataFrame) -> pd.DataFrame:
    existing = set(source_registry["Source_ID"].astype(str))
    add = [row for row in LOCAL_SOURCES if row["Source_ID"] not in existing]
    if add:
        source_registry = pd.concat([source_registry, pd.DataFrame(add)], ignore_index=True)
    return source_registry


def build_legacy_row_maps(sheets: dict[str, pd.DataFrame]) -> dict[str, dict[str, int]]:
    maps: dict[str, dict[str, int]] = {}
    for sheet_name, df in sheets.items():
        if not sheet_name.startswith("Legacy_") or "Original_Name" not in df.columns:
            continue
        row_map = {}
        for i, row in df.iterrows():
            name = str(row.get("Original_Name", "")).strip()
            if name and name not in row_map:
                row_map[name] = i + 2
        maps[sheet_name] = row_map
    return maps


def build_text_line_maps() -> dict[str, dict[str, int]]:
    files = {
        "Bennu": ROOT / "Python-MA/Bennu/Bennu.txt",
        "Ryugu": ROOT / "Python-MA/Bennu/Ryugu.txt",
        "PAH": ROOT / "Python-MA/Bennu/PAH.txt",
        "Biomarker": ROOT / "Python-MA/Bennu/Biomarker.txt",
        "Peptide": ROOT / "Python-MA/Bennu/Peptide.txt",
        "Prebio": ROOT / "Python-MA/prebio/prebio.txt",
    }
    result = {}
    for key, path in files.items():
        mapping = {}
        if not path.exists():
            result[key] = mapping
            continue
        for line_no, raw in enumerate(path.read_text(encoding="utf-8", errors="replace").splitlines(), start=1):
            line = raw.strip()
            if not line:
                continue
            # Prebio lines often start with an InChIKey followed by a name.
            parts = line.split(maxsplit=1)
            candidates = [line]
            if len(parts) == 2 and re.match(r"^[A-Z]{14}-[A-Z]{10}-[A-Z]$", parts[0]):
                candidates.append(parts[1].strip())
                candidates.append(parts[0].strip())
            for c in candidates:
                mapping.setdefault(normalize_name(c), line_no)
        result[key] = mapping
    return result


def build_prebio_condition_index() -> dict[str, list[dict]]:
    cond_dir = ROOT / "Python-MA/prebio/conditions"
    index: dict[str, list[dict]] = {}
    if not cond_dir.exists():
        return index
    for path in sorted(cond_dir.glob("pba-*.json")):
        try:
            data = json.loads(path.read_text(encoding="utf-8", errors="replace"))
        except Exception:
            continue
        blob = " ".join(
            str(data.get(k, ""))
            for k in ["primary", "smiles", "source", "key"]
        ).lower()
        entry = {
            "condition_id": data.get("key", path.stem),
            "file": str(path.relative_to(ROOT)),
            "source": str(data.get("source", "")).strip(),
            "primary": str(data.get("primary", "")).strip(),
        }
        for token in extract_condition_tokens(data):
            index.setdefault(normalize_name(token), []).append(entry)
        index.setdefault(blob, []).append(entry)
    return index


def extract_condition_tokens(data: dict) -> set[str]:
    tokens = set()
    for key in ["primary", "smiles"]:
        text = str(data.get(key, ""))
        for part in re.split(r"\s*\+\s*|\s*->\s*|>>|\.|,", text):
            part = part.strip()
            if part:
                tokens.add(part)
    return tokens


def normalize_name(value: object) -> str:
    if pd.isna(value):
        return ""
    return re.sub(r"[^a-z0-9]+", " ", str(value).lower()).strip()


def infer_source_location(
    origin: str,
    name: str,
    preferred: str,
    cur_row: pd.Series,
    all_row: int,
    legacy_row_maps: dict[str, dict[str, int]],
    text_line_maps: dict[str, dict[str, int]],
    prebio_conditions: dict[str, list[dict]],
) -> tuple[str, str, str]:
    local_bits = [f"Legacy_All row {all_row}"]
    source_id = ""
    status = "exact_local_position_primary_literature_pending"

    def legacy(sheet: str) -> None:
        row = legacy_row_maps.get(sheet, {}).get(name)
        if row:
            local_bits.append(f"{sheet} row {row}")

    def text_line(label: str, path: str, lookup_name: str | None = None) -> None:
        line = text_line_maps.get(label, {}).get(normalize_name(lookup_name or preferred))
        if not line and lookup_name is None:
            line = text_line_maps.get(label, {}).get(normalize_name(name))
        if line:
            local_bits.append(f"{path} line {line}")

    if origin == "Bennu":
        source_id = "SRC-0013"
        legacy("Legacy_Bennu")
        text_line("Bennu", "Python-MA/Bennu/Bennu.txt")
        local_bits.insert(0, "Literature source: SRC-0013 Glavin et al. 2025 Bennu returned-sample soluble organics; exact supplementary row still pending")
        status = "literature_source_linked_exact_local_row_primary_table_pending"
    elif origin == "Ryugu":
        source_id = "SRC-0048"
        legacy("Legacy_Ryugu")
        text_line("Ryugu", "Python-MA/Bennu/Ryugu.txt")
        local_bits.insert(0, "Literature source: SRC-0048 Naraoka et al. 2023 Ryugu soluble organic molecules; exact supplementary row still pending")
        status = "literature_source_linked_exact_local_row_primary_table_pending"
    elif origin == "Prebio":
        source_id, cond_bits = prebio_location(preferred, name, cur_row, prebio_conditions)
        legacy("Legacy_Prebio")
        text_line("Prebio", "Python-MA/prebio/prebio.txt")
        local_bits = cond_bits + local_bits
        status = "exact_reaction_condition_json" if cond_bits else "exact_local_row_primary_reaction_pending"
    elif origin == "PAHmethane pyrolysis":
        source_id = "SRC-0031"
        legacy("Legacy_methane pyrolysis")
        text_line("PAH", "Python-MA/Bennu/PAH.txt")
        local_bits.insert(0, "Literature source: SRC-0031 Khrabry et al. 2024 methane pyrolysis / PAH growth; species-label crosswalk still pending")
        status = "literature_source_linked_exact_local_row_species_crosswalk_pending"
    elif origin == "Combustion":
        source_id = "SRC-LOCAL-COMBUSTION-A"
        legacy("Legacy_Combustion")
        chem = ROOT / f"PAH ChemDraw/{name}.cdxml"
        if chem.exists():
            local_bits.append(f"PAH ChemDraw/{name}.cdxml")
        local_bits.insert(0, "Local source: combustion A-series curated structure set; primary combustion literature/table pending")
        status = "exact_local_row_primary_literature_pending"
    elif origin == "Biomarker":
        source_id = "SRC-0053"
        legacy("Legacy_Biomarker")
        text_line("Biomarker", "Python-MA/Bennu/Biomarker.txt")
        local_bits.insert(0, "Literature source: SRC-0053 Peters et al. Biomarker Guide; exact page/table pending")
        status = "literature_source_linked_exact_local_row_primary_page_pending"
    elif origin in {"peptide", "siderophore"}:
        source_id = "SRC-LOCAL-PEPTIDE"
        legacy("Legacy_Peptide")
        text_line("Peptide", "Python-MA/Bennu/Peptide.txt")
        local_bits.insert(0, "Local source: peptide/siderophore curated list; primary biological literature pending")
        status = "exact_local_row_primary_literature_pending"
    else:
        source_id = ""
        status = "unresolved"

    return source_id, " | ".join(dict.fromkeys([b for b in local_bits if b])), status


def prebio_location(preferred: str, name: str, cur_row: pd.Series, prebio_conditions: dict[str, list[dict]]) -> tuple[str, list[str]]:
    candidates = [
        normalize_name(preferred),
        normalize_name(name),
        normalize_name(cur_row.get("InChIKey", "")),
        normalize_name(cur_row.get("Canonical_SMILES", "")),
        normalize_name(cur_row.get("Formula", "")),
    ]
    matches = []
    for cand in candidates:
        if not cand:
            continue
        matches.extend(prebio_conditions.get(cand, []))
    # Fallback: substring search through blob keys.
    if not matches:
        for key, entries in prebio_conditions.items():
            if len(key) > 20:
                if normalize_name(preferred) and normalize_name(preferred) in key:
                    matches.extend(entries)
                elif normalize_name(name) and normalize_name(name) in key:
                    matches.extend(entries)
    unique = []
    seen = set()
    for m in matches:
        cid = m["condition_id"]
        if cid not in seen:
            unique.append(m)
            seen.add(cid)
    if not unique:
        return "SRC-0073", []
    source_ids = {"10.1038/s41586-019-1151-1": "SRC-0073"}
    first_source = unique[0].get("source", "").strip()
    source_id = source_ids.get(first_source, "SRC-0073")
    ids = ", ".join(m["condition_id"] for m in unique[:8])
    files = ", ".join(m["file"] for m in unique[:3])
    doi = first_source or "source DOI not recorded"
    return source_id, [f"Prebiotic reaction condition(s): {ids}; files: {files}; source DOI: {doi}"]


def rebuild_source_link_queue(evidence: pd.DataFrame, curated: pd.DataFrame) -> pd.DataFrame:
    missing_source = evidence["Source_ID"].isna() | evidence["Source_ID"].astype(str).str.strip().eq("")
    missing_location = evidence["Evidence_Location"].isna() | evidence["Evidence_Location"].astype(str).str.strip().eq("")
    ev = evidence[missing_source | missing_location].copy()
    cur_cols = ["Record_IDs", "Curated_ID", "Preferred_Name", "Origin", "MA", "Confidence"]
    ev = ev.merge(curated[cur_cols], left_on="Record_ID", right_on="Record_IDs", how="left")
    if ev.empty:
        return pd.DataFrame(columns=["Curated_ID", "Record_ID", "Preferred_Name", "Origin", "MA", "Source_ID", "Evidence_Location", "Source_Task"])
    ev["Source_Task"] = "link_source_id_or_evidence_location"
    cols = ["Curated_ID", "Record_ID", "Preferred_Name", "Origin", "MA", "Source_ID", "Evidence_Location", "Source_Task", "Evidence_Confidence", "Notes"]
    return ev[[c for c in cols if c in ev.columns]]


def build_primary_literature_audit_queue(evidence: pd.DataFrame, curated: pd.DataFrame) -> pd.DataFrame:
    mask = evidence["Evidence_Position_Status"].fillna("").astype(str).str.contains("pending|local", case=False, regex=True)
    ev = evidence[mask].copy()
    ev = ev.merge(curated[["Record_IDs", "Curated_ID", "Preferred_Name", "Origin", "MA"]], left_on="Record_ID", right_on="Record_IDs", how="left")
    ev["Audit_Task"] = ev["Evidence_Position_Status"].map(audit_task)
    cols = ["Curated_ID", "Record_ID", "Preferred_Name", "Origin", "MA", "Source_ID", "Evidence_Position_Status", "Evidence_Location", "Audit_Task"]
    return ev[[c for c in cols if c in ev.columns]]


def audit_task(status: str) -> str:
    if "primary_table_pending" in str(status):
        return "find_exact_primary_paper_supplementary_table_row"
    if "species_crosswalk_pending" in str(status):
        return "map_local_species_name_to_published_mechanism_species_label"
    if "primary_page_pending" in str(status):
        return "find_exact_book_page_or_table"
    if "primary_literature_pending" in str(status):
        return "identify_primary_literature_source"
    return "audit_source_position"


def update_dictionary(dictionary: pd.DataFrame) -> pd.DataFrame:
    additions = [
        ("Evidence_Log", "Evidence_Position_Status", "Whether Evidence_Location is exact local position, reaction JSON, or still needs primary literature table/page audit."),
        ("Source_Position_Update_Log", "Evidence_Location", "Location string assigned during v0.6 source-location pass."),
        ("Primary_Literature_Audit_Queue", "Audit_Task", "Remaining task needed to convert local/source-domain provenance into publication-grade exact source location."),
    ]
    add_df = pd.DataFrame(additions, columns=["Sheet", "Field", "Meaning"])
    merged = pd.concat([dictionary, add_df], ignore_index=True)
    return merged.drop_duplicates(subset=["Sheet", "Field"], keep="last")


def update_checks(checks: pd.DataFrame, evidence: pd.DataFrame, sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    missing_source = int((evidence["Source_ID"].isna() | evidence["Source_ID"].astype(str).str.strip().eq("")).sum())
    missing_location = int((evidence["Evidence_Location"].isna() | evidence["Evidence_Location"].astype(str).str.strip().eq("")).sum())
    additions = pd.DataFrame(
        [
            {"Check": "v0.6 Evidence_Log missing Source_ID", "Value": missing_source},
            {"Check": "v0.6 Evidence_Log missing Evidence_Location", "Value": missing_location},
            {"Check": "v0.6 source-position update rows", "Value": len(sheets["Source_Position_Update_Log"])},
            {"Check": "v0.6 primary literature audit rows", "Value": len(sheets["Primary_Literature_Audit_Queue"])},
        ]
    )
    return pd.concat([checks, additions], ignore_index=True)


def update_changelog(changelog: pd.DataFrame) -> pd.DataFrame:
    row = {
        "Version": "v0.6_source_locations",
        "Date": "2026-06-11",
        "Input_Workbook": str(INPUT.relative_to(ROOT)),
        "Input_Bib": "数据库/Assembly Method.bib",
        "Input_Workbook_SHA256": "",
        "Input_Bib_SHA256": "",
        "Major_Changes": "Filled Source_ID and Evidence_Location with exact local workbook/list/reaction-condition positions where available; added Source_Position_Update_Log and Primary_Literature_Audit_Queue.",
    }
    return pd.concat([changelog, pd.DataFrame([row])], ignore_index=True)


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    preferred_order = [
        "README",
        "Change_Log",
        "Data_Dictionary",
        "Curation_Priority",
        "Source_Position_Update_Log",
        "Primary_Literature_Audit_Queue",
        "MA_Recompute_Queue",
        "Structure_Audit_Log",
        "Nonbiotic_Search_Log",
        "Source_Link_Task_Queue",
        "MA_Removal_Log",
        "Pilot_Curation_20",
        "Source_Registry",
        "Source_Aliases",
        "Evidence_Log",
        "Molecule_Curated",
        "ML_Dataset_View",
        "High_MA_Abiotic",
        "Low_MA_Bio",
        "Curation_Checks",
    ]
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for name in preferred_order:
            if name in sheets:
                sheets[name].to_excel(writer, sheet_name=name[:31], index=False)
        for name, df in sheets.items():
            if name not in preferred_order:
                df.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(OUTPUT)
    fills = {
        "Source_Position_Update_Log": "5B9BD5",
        "Primary_Literature_Audit_Queue": "C55A11",
        "Source_Link_Task_Queue": "1F4E78",
    }
    default_fill = "1F4E78"
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = PatternFill("solid", fgColor=fills.get(ws.title, default_fill))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, min(ws.max_column, 70) + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter][: min(ws.max_row, 160)]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 100))
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 60))
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
