from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "数据库"
INPUT = DATA_DIR / "5.20完善数据集_research_db_v0.4_ma_corrected.xlsx"
OUTPUT = DATA_DIR / "5.20完善数据集_research_db_v0.5_workflow_enhanced.xlsx"

ASSEMBLY_URL = "https://github.com/DaymudeLab/assembly-theory"
ASSEMBLY_IMPL = "DaymudeLab/assembly-theory"


MA_PROVENANCE_FIELDS = [
    "MA_Method",
    "MA_Implementation",
    "MA_Code_URL",
    "MA_Code_Commit",
    "MA_Input_File",
    "MA_Input_SHA256",
    "MA_Run_Date",
    "MA_Status",
    "MA_Runtime_or_Timeout",
    "MA_Error_or_Note",
]


def main() -> None:
    sheets = pd.read_excel(INPUT, sheet_name=None)
    curated = sheets["Molecule_Curated"].copy()
    removal = sheets.get("MA_Removal_Log", pd.DataFrame())

    curated = add_ma_provenance(curated, removal)
    sheets["Molecule_Curated"] = curated
    sheets["ML_Dataset_View"] = update_ml_view(sheets["ML_Dataset_View"].copy(), curated)
    sheets["MA_Recompute_Queue"] = build_ma_recompute_queue(curated, removal)
    sheets["Structure_Audit_Log"] = build_structure_audit_log(curated)
    sheets["Nonbiotic_Search_Log"] = build_nonbiotic_search_log(curated)
    sheets["Source_Link_Task_Queue"] = build_source_link_queue(curated, sheets["Evidence_Log"])
    sheets["Curation_Priority"] = build_priority_sheet(
        sheets["MA_Recompute_Queue"],
        sheets["Structure_Audit_Log"],
        sheets["Nonbiotic_Search_Log"],
        sheets["Source_Link_Task_Queue"],
    )
    sheets["Data_Dictionary"] = update_dictionary(sheets["Data_Dictionary"])
    sheets["Curation_Checks"] = update_checks(sheets["Curation_Checks"], sheets)
    sheets["Change_Log"] = update_changelog(sheets["Change_Log"])

    write_workbook(sheets)
    print(OUTPUT)


def add_ma_provenance(curated: pd.DataFrame, removal: pd.DataFrame) -> pd.DataFrame:
    for field in MA_PROVENANCE_FIELDS:
        if field not in curated.columns:
            curated[field] = ""

    removed_records = set(removal["Record_ID"]) if not removal.empty else set()
    ma_present = curated["MA"].notna()
    ma_missing = ~ma_present
    removed = curated["Record_IDs"].isin(removed_records)

    curated.loc[ma_present, "MA_Method"] = "Molecular Assembly Index"
    curated.loc[ma_present, "MA_Implementation"] = ASSEMBLY_IMPL
    curated.loc[ma_present, "MA_Code_URL"] = ASSEMBLY_URL
    curated.loc[ma_present, "MA_Status"] = "legacy_computed_needs_provenance"
    curated.loc[ma_present, "MA_Error_or_Note"] = (
        "MA value exists in legacy workbook; code commit, input mol file, and run log not yet recorded."
    )

    curated.loc[ma_missing, "MA_Method"] = "Molecular Assembly Index"
    curated.loc[ma_missing, "MA_Implementation"] = ASSEMBLY_IMPL
    curated.loc[ma_missing, "MA_Code_URL"] = ASSEMBLY_URL
    curated.loc[ma_missing, "MA_Status"] = "missing_needs_recompute"
    curated.loc[ma_missing, "MA_Error_or_Note"] = "MA missing; compute with auditable mol input and assembly-theory version."

    curated.loc[removed, "MA_Status"] = "removed_random_placeholder_needs_recompute"
    curated.loc[removed, "MA_Error_or_Note"] = (
        "Original MA was yellow-highlighted/random placeholder and removed in v0.4; recompute before analysis."
    )

    # If a mol file is already listed, keep it as the intended MA input. Otherwise leave explicit blank.
    has_mol = curated["Mol_File"].notna() & curated["Mol_File"].astype(str).str.strip().ne("")
    curated.loc[has_mol, "MA_Input_File"] = curated.loc[has_mol, "Mol_File"]
    curated.loc[has_mol, "MA_Input_SHA256"] = curated.loc[has_mol, "Mol_File_SHA256"]
    return curated


def update_ml_view(ml: pd.DataFrame, curated: pd.DataFrame) -> pd.DataFrame:
    keep = [
        "Curated_ID",
        "MA_Method",
        "MA_Implementation",
        "MA_Code_URL",
        "MA_Code_Commit",
        "MA_Input_File",
        "MA_Input_SHA256",
        "MA_Run_Date",
        "MA_Status",
        "MA_Runtime_or_Timeout",
        "MA_Error_or_Note",
    ]
    merged = ml.drop(columns=[c for c in keep[1:] if c in ml.columns]).merge(curated[keep], on="Curated_ID", how="left")
    invalid_ma = merged["MA_Status"].isin(["missing_needs_recompute", "removed_random_placeholder_needs_recompute"])
    merged.loc[invalid_ma, "Train_Eligible"] = "no"
    merged.loc[invalid_ma, "Reason_Excluded"] = (
        "MA missing or removed; recompute with recorded assembly-theory version and mol input before ML use"
    )
    return merged


def build_ma_recompute_queue(curated: pd.DataFrame, removal: pd.DataFrame) -> pd.DataFrame:
    needs = curated[
        curated["MA_Status"].isin(["missing_needs_recompute", "removed_random_placeholder_needs_recompute"])
        | curated["MA_Code_Commit"].fillna("").astype(str).str.strip().eq("")
    ].copy()

    # Prioritize removed/random rows first, then rows with existing MA but missing provenance.
    def priority(row: pd.Series) -> str:
        if row["MA_Status"] == "removed_random_placeholder_needs_recompute":
            return "P0_recompute_removed_random_MA"
        if pd.isna(row["MA"]):
            return "P1_compute_missing_MA"
        return "P2_add_provenance_or_recompute"

    needs["Task_Priority"] = needs.apply(priority, axis=1)
    needs["Recommended_Input"] = needs.apply(
        lambda r: r["Mol_File"] if pd.notna(r["Mol_File"]) and str(r["Mol_File"]).strip() else "create_or_link_mol_file",
        axis=1,
    )
    needs["Required_Action"] = needs.apply(
        lambda r: "recompute_with_assembly_theory" if r["Task_Priority"].startswith(("P0", "P1")) else "record_code_commit_input_hash_run_log",
        axis=1,
    )
    cols = [
        "Curated_ID",
        "Record_IDs",
        "Preferred_Name",
        "Origin",
        "Formula",
        "Old_or_Current_MA",
        "MA_Status",
        "Task_Priority",
        "Recommended_Input",
        "MA_Implementation",
        "MA_Code_URL",
        "MA_Code_Commit",
        "MA_Input_File",
        "MA_Input_SHA256",
        "Required_Action",
        "MA_Error_or_Note",
    ]
    needs["Old_or_Current_MA"] = needs["MA"]
    return needs[[c for c in cols if c in needs.columns]].sort_values(["Task_Priority", "Origin", "Preferred_Name"])


def build_structure_audit_log(curated: pd.DataFrame) -> pd.DataFrame:
    mask = (
        curated["Canonical_SMILES"].isna()
        | curated["InChIKey"].isna()
        | curated["Mol_File"].isna()
        | curated["Mol_File"].astype(str).str.strip().eq("")
        | curated["Graph_Level"].astype(str).str.startswith("G0")
    )
    df = curated[mask].copy()
    df["Structure_Audit_Status"] = "not_started"
    df["Structure_Audit_Task"] = df.apply(structure_task, axis=1)
    cols = [
        "Curated_ID",
        "Record_IDs",
        "Preferred_Name",
        "Origin",
        "Formula",
        "Canonical_SMILES",
        "InChIKey",
        "Mol_File",
        "Mol_File_SHA256",
        "Graph_Level",
        "Graph_Ambiguity",
        "Structure_Validation_Status",
        "Structure_Audit_Task",
        "Structure_Audit_Status",
    ]
    return df[[c for c in cols if c in df.columns]]


def structure_task(row: pd.Series) -> str:
    if str(row.get("Graph_Level", "")).startswith("G0"):
        return "resolve_single_molecular_graph_or_mark_context_only"
    if pd.isna(row.get("Mol_File")) or str(row.get("Mol_File", "")).strip() == "":
        return "link_or_generate_mol_file_and_sha256"
    if pd.isna(row.get("Canonical_SMILES")) or pd.isna(row.get("InChIKey")):
        return "derive_standard_identifiers_from_mol"
    return "audit_structure_fields"


def build_nonbiotic_search_log(curated: pd.DataFrame) -> pd.DataFrame:
    bio = curated[curated["Origin_Label"].eq("biological")].copy()
    bio["Search_Status"] = "not_started"
    bio["Search_Query_Template"] = bio["Preferred_Name"].map(
        lambda x: f'"{x}" abiotic OR nonbiological OR prebiotic OR meteorite OR combustion OR hydrothermal'
    )
    bio["Search_Databases"] = "Google Scholar; Web of Science; PubChem; Reaxys/SciFinder if available"
    bio["Decision_Rule"] = (
        "If reliable nonbiotic formation/detection is found, downgrade from Bio_Strong_Candidate to Bio_Ambiguous or mark conflict."
    )
    cols = [
        "Curated_ID",
        "Preferred_Name",
        "Origin",
        "MA",
        "Origin_Class",
        "Label_Strength",
        "Nonbiotic_Report_Status",
        "Search_Status",
        "Search_Query_Template",
        "Search_Databases",
        "Decision_Rule",
    ]
    return bio[[c for c in cols if c in bio.columns]].sort_values(["Origin_Class", "Preferred_Name"])


def build_source_link_queue(curated: pd.DataFrame, evidence: pd.DataFrame) -> pd.DataFrame:
    missing_source = evidence["Source_ID"].isna() | evidence["Source_ID"].astype(str).str.strip().eq("")
    missing_location = evidence["Evidence_Location"].isna() | evidence["Evidence_Location"].astype(str).str.strip().eq("")
    ev = evidence[missing_source | missing_location].copy()
    cur_cols = ["Record_IDs", "Curated_ID", "Preferred_Name", "Origin", "MA", "Confidence"]
    ev = ev.merge(curated[cur_cols], left_on="Record_ID", right_on="Record_IDs", how="left")
    ev["Source_Task"] = ev.apply(
        lambda r: "link_source_id_and_evidence_location"
        if pd.isna(r.get("Source_ID")) or str(r.get("Source_ID", "")).strip() == ""
        else "fill_exact_evidence_location",
        axis=1,
    )
    cols = [
        "Curated_ID",
        "Record_ID",
        "Preferred_Name",
        "Origin",
        "MA",
        "Source_ID",
        "Evidence_Location",
        "Source_Task",
        "Evidence_Confidence",
        "Notes",
    ]
    return ev[[c for c in cols if c in ev.columns]]


def build_priority_sheet(ma_queue: pd.DataFrame, structure_log: pd.DataFrame, nonbiotic_log: pd.DataFrame, source_queue: pd.DataFrame) -> pd.DataFrame:
    rows = [
        {
            "Priority": "P0",
            "Task": "Recompute removed/random MA values",
            "Sheet": "MA_Recompute_Queue",
            "Rows": int(ma_queue["Task_Priority"].astype(str).str.startswith("P0").sum()) if not ma_queue.empty else 0,
            "Why": "These values were random placeholders and must not support MA conclusions.",
        },
        {
            "Priority": "P1",
            "Task": "Add mol files and SHA256 hashes",
            "Sheet": "Structure_Audit_Log",
            "Rows": len(structure_log),
            "Why": "MA reproducibility requires auditable molecular graph inputs.",
        },
        {
            "Priority": "P2",
            "Task": "Link exact source tables/figures",
            "Sheet": "Source_Link_Task_Queue",
            "Rows": len(source_queue),
            "Why": "Claims must trace back to exact literature/database evidence.",
        },
        {
            "Priority": "P3",
            "Task": "Run nonbiotic-origin searches for biological candidates",
            "Sheet": "Nonbiotic_Search_Log",
            "Rows": len(nonbiotic_log),
            "Why": "Biomarker labels are not life-exclusive until nonbiotic reports are checked.",
        },
    ]
    return pd.DataFrame(rows)


def update_dictionary(dictionary: pd.DataFrame) -> pd.DataFrame:
    additions = [
        ("Molecule_Curated", "MA_Method", "Name of the molecular complexity metric; here Molecular Assembly Index."),
        ("Molecule_Curated", "MA_Implementation", "Software implementation used for MA; current source is DaymudeLab/assembly-theory."),
        ("Molecule_Curated", "MA_Code_Commit", "Exact git commit or version of the MA code used for this row."),
        ("Molecule_Curated", "MA_Input_File", "Path to the mol/sdf or other molecular graph input used for MA calculation."),
        ("Molecule_Curated", "MA_Input_SHA256", "Hash of the MA input file for reproducibility."),
        ("Molecule_Curated", "MA_Status", "Computed/provenance/recompute status of the MA value."),
        ("MA_Recompute_Queue", "Task_Priority", "Priority for MA recomputation or provenance completion."),
        ("Structure_Audit_Log", "Structure_Audit_Task", "Required action to make the molecular graph auditable."),
        ("Nonbiotic_Search_Log", "Search_Query_Template", "Suggested query for checking nonbiotic reports of biological candidates."),
        ("Source_Link_Task_Queue", "Source_Task", "Required source-linking action for evidence records."),
    ]
    add_df = pd.DataFrame(additions, columns=["Sheet", "Field", "Meaning"])
    merged = pd.concat([dictionary, add_df], ignore_index=True)
    return merged.drop_duplicates(subset=["Sheet", "Field"], keep="last")


def update_checks(checks: pd.DataFrame, sheets: dict[str, pd.DataFrame]) -> pd.DataFrame:
    curated = sheets["Molecule_Curated"]
    additions = pd.DataFrame(
        [
            {"Check": "v0.5 current curated rows", "Value": len(curated)},
            {"Check": "v0.5 current MA missing", "Value": int(curated["MA"].isna().sum())},
            {"Check": "v0.5 current MA with legacy/provenance gap", "Value": int(curated["MA_Status"].eq("legacy_computed_needs_provenance").sum())},
            {"Check": "v0.5 MA recompute/provenance queue rows", "Value": len(sheets["MA_Recompute_Queue"])},
            {"Check": "v0.5 P0 removed/random MA recompute rows", "Value": int(sheets["MA_Recompute_Queue"]["Task_Priority"].astype(str).str.startswith("P0").sum())},
            {"Check": "v0.5 structure audit rows", "Value": len(sheets["Structure_Audit_Log"])},
            {"Check": "v0.5 nonbiotic search rows", "Value": len(sheets["Nonbiotic_Search_Log"])},
            {"Check": "v0.5 source link task rows", "Value": len(sheets["Source_Link_Task_Queue"])},
        ]
    )
    return pd.concat([checks, additions], ignore_index=True)


def update_changelog(changelog: pd.DataFrame) -> pd.DataFrame:
    row = {
        "Version": "v0.5_workflow_enhanced",
        "Date": "2026-06-11",
        "Input_Workbook": str(INPUT.relative_to(ROOT)),
        "Input_Bib": "数据库/Assembly Method.bib",
        "Input_Workbook_SHA256": "",
        "Input_Bib_SHA256": "",
        "Major_Changes": "Added MA provenance fields, MA_Recompute_Queue, Structure_Audit_Log, Nonbiotic_Search_Log, Source_Link_Task_Queue, and Curation_Priority.",
    }
    return pd.concat([changelog, pd.DataFrame([row])], ignore_index=True)


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    preferred_order = [
        "README",
        "Change_Log",
        "Data_Dictionary",
        "Curation_Priority",
        "MA_Recompute_Queue",
        "Structure_Audit_Log",
        "Nonbiotic_Search_Log",
        "Source_Link_Task_Queue",
        "MA_Removal_Log",
        "Pilot_Curation_20",
        "Source_Registry",
        "Source_Aliases",
        "Evidence_Log",
        "Molecule_Curated",
        "ML_Dataset_View",
        "High_MA_Abiotic",
        "Low_MA_Bio",
        "Curation_Checks",
    ]
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for name in preferred_order:
            if name in sheets:
                sheets[name].to_excel(writer, sheet_name=name[:31], index=False)
        for name, df in sheets.items():
            if name not in preferred_order:
                df.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(OUTPUT)
    fills = {
        "Curation_Priority": "7030A0",
        "MA_Recompute_Queue": "C00000",
        "Structure_Audit_Log": "9E480E",
        "Nonbiotic_Search_Log": "548235",
        "Source_Link_Task_Queue": "1F4E78",
    }
    default_fill = "1F4E78"
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = PatternFill("solid", fgColor=fills.get(ws.title, default_fill))
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, min(ws.max_column, 70) + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter][: min(ws.max_row, 160)]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 90))
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 55))
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
