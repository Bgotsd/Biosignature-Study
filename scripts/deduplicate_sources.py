from __future__ import annotations

import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "数据库"
INPUT = DATA_DIR / "5.20完善数据集_research_db_v0.2_pilot20.xlsx"
OUTPUT = DATA_DIR / "5.20完善数据集_research_db_v0.3_sources_dedup.xlsx"


def norm_doi(value: object) -> str:
    if pd.isna(value):
        return ""
    s = str(value).strip().lower()
    s = re.sub(r"^https?://(dx\.)?doi\.org/", "", s)
    return s


def norm_title(value: object) -> str:
    if pd.isna(value):
        return ""
    s = str(value).lower()
    return re.sub(r"[^a-z0-9]+", " ", s).strip()


def canonical_key(row: pd.Series) -> str:
    doi = row["Normalized_DOI"]
    if doi:
        return "doi:" + doi
    title = row["Normalized_Title"]
    if title:
        return "title:" + title
    return "source_id:" + str(row["Source_ID"])


def choose_canonical(group: pd.DataFrame) -> str:
    # Prefer rows with DOI, year, journal, and no suffix-like trailing "a/b" key.
    scored = []
    for _, row in group.iterrows():
        key = str(row.get("BibTeX_Key", ""))
        score = 0
        if norm_doi(row.get("DOI_or_URL")):
            score += 8
        if pd.notna(row.get("Year")):
            score += 3
        if pd.notna(row.get("Journal")) and str(row.get("Journal")).strip():
            score += 2
        if not re.search(r"\d{4}[a-z]$", key):
            score += 1
        # Keep stable order as tie-breaker by lower numeric Source_ID.
        sid_num = int(re.sub(r"\D", "", str(row["Source_ID"])) or 999999)
        scored.append((score, -sid_num, row["Source_ID"]))
    scored.sort(reverse=True)
    return scored[0][2]


def main() -> None:
    sheets = pd.read_excel(INPUT, sheet_name=None)
    src = sheets["Source_Registry"].copy()
    src["Normalized_DOI"] = src["DOI_or_URL"].map(norm_doi)
    src["Normalized_Title"] = src["Title"].map(norm_title)
    src["Duplicate_Key"] = src.apply(canonical_key, axis=1)

    canonical_map: dict[str, str] = {}
    duplicate_group_map: dict[str, str] = {}
    alias_rows = []
    canonical_rows = []

    for group_idx, (dup_key, group) in enumerate(src.groupby("Duplicate_Key", sort=False), start=1):
        canonical_id = choose_canonical(group)
        group_id = f"DUP-{group_idx:04d}" if len(group) > 1 else ""
        for _, row in group.iterrows():
            canonical_map[row["Source_ID"]] = canonical_id
            duplicate_group_map[row["Source_ID"]] = group_id
            alias_rows.append(
                {
                    "Original_Source_ID": row["Source_ID"],
                    "Canonical_Source_ID": canonical_id,
                    "Duplicate_Group": group_id,
                    "Duplicate_Key": dup_key,
                    "BibTeX_Key": row.get("BibTeX_Key", ""),
                    "Title": row.get("Title", ""),
                    "DOI_or_URL": row.get("DOI_or_URL", ""),
                    "Alias_Status": "canonical" if row["Source_ID"] == canonical_id else "duplicate_alias",
                }
            )
        canonical = group[group["Source_ID"].eq(canonical_id)].iloc[0].copy()
        canonical["Alias_Source_IDs"] = "; ".join(group["Source_ID"].astype(str).tolist())
        canonical["Alias_BibTeX_Keys"] = "; ".join(group["BibTeX_Key"].fillna("").astype(str).tolist())
        canonical["Duplicate_Group"] = group_id
        canonical["Duplicate_Count"] = len(group)
        canonical_rows.append(canonical)

    canonical_src = pd.DataFrame(canonical_rows)
    aliases = pd.DataFrame(alias_rows)

    # Preserve Source_ID as canonical IDs, and update references in evidence/pilot.
    for sheet_name in ["Evidence_Log", "Pilot_Curation_20"]:
        if sheet_name in sheets and "Source_ID" in sheets[sheet_name].columns:
            df = sheets[sheet_name].copy()
            df["Original_Source_ID_Before_Dedup"] = df["Source_ID"]
            df["Source_ID"] = df["Source_ID"].map(lambda x: canonical_map.get(x, x))
            sheets[sheet_name] = df

    # Add canonical fields to source registry.
    source_cols = [
        "Source_ID",
        "BibTeX_Key",
        "Source_Type",
        "Title",
        "Authors",
        "Year",
        "Journal",
        "DOI_or_URL",
        "PMID",
        "Source_Domain",
        "Search_Query",
        "Why_Relevant",
        "Access_Date",
        "Screening_Status",
        "Exclusion_Reason",
        "Notes",
        "Normalized_DOI",
        "Normalized_Title",
        "Duplicate_Key",
        "Duplicate_Group",
        "Duplicate_Count",
        "Alias_Source_IDs",
        "Alias_BibTeX_Keys",
    ]
    canonical_src = canonical_src[[c for c in source_cols if c in canonical_src.columns]]
    sheets["Source_Registry"] = canonical_src
    sheets["Source_Aliases"] = aliases

    checks = sheets["Curation_Checks"].copy()
    duplicate_alias_count = int((aliases["Alias_Status"] == "duplicate_alias").sum())
    extra = pd.DataFrame(
        [
            {"Check": "Canonical source rows after dedup", "Value": len(canonical_src)},
            {"Check": "Duplicate source aliases", "Value": duplicate_alias_count},
            {
                "Check": "Duplicate source groups",
                "Value": int(aliases.loc[aliases["Duplicate_Group"].notna() & aliases["Duplicate_Group"].ne(""), "Duplicate_Group"].nunique()),
            },
            {
                "Check": "Source rows belonging to duplicate groups",
                "Value": int(aliases["Duplicate_Group"].fillna("").astype(str).str.startswith("DUP-").sum()),
            },
        ]
    )
    sheets["Curation_Checks"] = pd.concat([checks, extra], ignore_index=True)

    changelog = sheets["Change_Log"].copy()
    changelog = pd.concat(
        [
            changelog,
            pd.DataFrame(
                [
                    {
                        "Version": "v0.3_sources_dedup",
                        "Date": "2026-06-11",
                        "Input_Workbook": str(INPUT.relative_to(ROOT)),
                        "Input_Bib": "数据库/Assembly Method.bib",
                        "Input_Workbook_SHA256": "",
                        "Input_Bib_SHA256": "",
                        "Major_Changes": "Deduplicated Source_Registry by normalized DOI/title; added Source_Aliases; updated Evidence_Log and Pilot_Curation_20 to canonical Source_IDs.",
                    }
                ]
            ),
        ],
        ignore_index=True,
    )
    sheets["Change_Log"] = changelog

    write_workbook(sheets)
    print(OUTPUT)


def write_workbook(sheets: dict[str, pd.DataFrame]) -> None:
    preferred_order = [
        "README",
        "Change_Log",
        "Data_Dictionary",
        "Pilot_Curation_20",
        "Source_Registry",
        "Source_Aliases",
        "Evidence_Log",
        "Molecule_Curated",
        "ML_Dataset_View",
        "High_MA_Abiotic",
        "Low_MA_Bio",
        "Curation_Checks",
    ]
    with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:
        for name in preferred_order:
            if name in sheets:
                sheets[name].to_excel(writer, sheet_name=name[:31], index=False)
        for name, df in sheets.items():
            if name not in preferred_order:
                df.to_excel(writer, sheet_name=name[:31], index=False)

    wb = load_workbook(OUTPUT)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    alias_fill = PatternFill("solid", fgColor="7F6000")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        fill = alias_fill if ws.title == "Source_Aliases" else header_fill
        for cell in ws[1]:
            cell.fill = fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col_idx in range(1, min(ws.max_column, 55) + 1):
            letter = get_column_letter(col_idx)
            max_len = 0
            for cell in ws[letter][: min(ws.max_row, 150)]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 90))
            ws.column_dimensions[letter].width = max(12, min(max_len + 2, 55))
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
